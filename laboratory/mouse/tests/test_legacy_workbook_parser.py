from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "parse_legacy_workbooks.py"


def load_parser_module():
    spec = importlib.util.spec_from_file_location("parse_legacy_workbooks", SCRIPT_PATH)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_parse_animal_sheet_keeps_workbook_rows_reviewable(tmp_path: Path) -> None:
    parser = load_parser_module()
    workbook_path = tmp_path / "animal-sheet.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "animal sheet"
    ws.append(["Cage No.", "Strain", "Sex", "I.D", "Genotype", "DOB", "Mating Date", "Pubs"])
    ws.append(["C-014", "ApoM Tg/Tg", "\u2642", "319", "Tg/Tg", "2026-04-13", "", ""])
    ws.append(["", "", "", "", "", "", "2026-04-13", "10p"])
    wb.save(workbook_path)

    payload = parser.parse_workbook(workbook_path, kind="animal")

    assert payload["layer"] == "parsed or intermediate result"
    assert payload["source_layer"] == "export or view"
    assert payload["workbook_kind"] == "legacy_animal_sheet"
    assert payload["notes"]
    assert payload["rows"][0]["row_type"] == "parent_or_mouse_snapshot"
    assert payload["rows"][0]["sex_raw"] == "\u2642"
    assert payload["rows"][0]["sex_candidate"] == "male"
    assert payload["rows"][0]["cage_no_raw"] == "C-014"
    assert payload["rows"][0]["source_cells"]["display_id"] == "D2"
    assert payload["rows"][1]["row_type"] == "litter_or_offspring_snapshot"
    assert payload["rows"][1]["review_status"] == "candidate"


def test_parse_separation_sheet_keeps_counts_as_candidates(tmp_path: Path) -> None:
    parser = load_parser_module()
    workbook_path = tmp_path / "separation.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "separation"
    ws.append(["Strain", "Genotype", "Total", "DOB", "WT", "TG", "Sampling Point"])
    ws.append(["Npc1 fl/fl", "fl/fl", "\u2640 3p", "2026-04-01", "1", "2", "tail"])
    wb.save(workbook_path)

    payload = parser.parse_workbook(workbook_path, kind="separation")

    row = payload["rows"][0]
    assert payload["workbook_kind"] == "legacy_separation_status"
    assert row["row_type"] == "separation_summary_row"
    assert row["source_layer"] == "export or view"
    assert row["sex_candidate"] == "female"
    assert row["count_candidate"] == 3
    assert row["source_cells"]["total"] == "C2"
    assert row["review_status"] == "candidate"


def test_parse_multisheet_animal_workbook_with_legacy_headers(tmp_path: Path) -> None:
    parser = load_parser_module()
    workbook_path = tmp_path / "legacy-animalsheet.xlsx"
    wb = Workbook()
    first = wb.active
    first.title = "GFAP Cre; S1PR1 flox"
    first.append(["Cage No.", "sex", "I.D", "", "DOB", "Mating date", "pup", "place"])
    first.append(["GFAP Cre; S1PR1 fl/fl", "\u2640", "GS", "cre; flox", "2023-03-28~04-14", "2024-02-01", "", "rack"])
    second = wb.create_sheet("ptgs2 S565A flox")
    second.append(["Cage No.", "sex", "I.D", "I.D", "DOB", "Mating date", "pup", "place"])
    second.append(["ptgs2 S565A fl/fl", "F1", "16p", "separated", "2023-10-13", "", "", "rack"])
    wb.save(workbook_path)

    payload = parser.parse_workbook(workbook_path, kind="animal")

    assert payload["workbook_kind"] == "legacy_animal_sheet"
    assert "GFAP Cre; S1PR1 flox" in payload["sheet_name"]
    assert "ptgs2 S565A flox" in payload["sheet_name"]
    assert len(payload["rows"]) == 2
    assert payload["rows"][0]["strain_raw"] == "GFAP Cre; S1PR1 fl/fl"
    assert payload["rows"][0]["genotype_raw"] == "cre; flox"
    assert payload["rows"][0]["source_cells"]["genotype"] == "D2"
    assert payload["rows"][1]["row_type"] == "litter_or_offspring_snapshot"


def test_parse_multisheet_separation_workbook_with_sex_column(tmp_path: Path) -> None:
    parser = load_parser_module()
    workbook_path = tmp_path / "legacy-separation.xlsx"
    wb = Workbook()
    first = wb.active
    first.title = "Person A"
    first.append(["Cage number", "Strain", "Genotype", "SEX", "DOB", "Use"])
    first.append(["1", "TASTPMSgpl1+/-", "TASTPMSgpl1+/-", "\u2642-2p", "2023.06.09~06.30", "mating"])
    second = wb.create_sheet("Person B")
    second.append(["Cage number", "Strain", "Genotype", "Sex", "DOB", "Use"])
    second.append(["1", "ACfl/fl", "ACfl/fl", "\u2640-9p", "23.07.07~14", "mating"])
    wb.save(workbook_path)

    payload = parser.parse_workbook(workbook_path, kind="separation")

    assert payload["workbook_kind"] == "legacy_separation_status"
    assert len(payload["rows"]) == 2
    assert payload["rows"][0]["sex_candidate"] == "male"
    assert payload["rows"][0]["count_candidate"] == 2
    assert payload["rows"][0]["source_cells"]["total"] == "D2"
    assert payload["rows"][1]["sex_candidate"] == "female"
    assert payload["rows"][1]["count_candidate"] == 9
