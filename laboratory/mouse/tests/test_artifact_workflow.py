from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import db
from app import main as app_main


def test_persist_proposed_changeset_artifact_keeps_preview_non_canonical(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    preview = {
        "candidate_id": "candidate_001",
        "review_id": "review_001",
        "parse_id": "parse_001",
        "legacy_row_id": "legacy_001",
        "proposed_mice": [
            {
                "mouse_id": "mouse_MT318_parse_001",
                "display_id": "MT318",
                "source_note_item_id": "note_001",
                "source_photo_id": "photo_001",
                "card_snapshot_id": "card_001",
                "will_create_mouse": True,
                "will_create_event": True,
            }
        ],
        "duplicate_risks": [],
        "blocked": False,
        "blockers": [],
    }

    result = app_main.persist_proposed_changeset_artifact(
        preview,
        created_at="2026-05-09T12:00:00Z",
    )

    artifact_path = Path(result["artifact_path"])
    assert artifact_path.exists()
    assert artifact_path.parent == tmp_path / "mousedb_artifacts" / "proposed_changesets"

    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
    assert artifact["artifact_type"] == "proposed_changeset"
    assert artifact["source_layer"] == "export or view"
    assert artifact["status"] == "draft"
    assert artifact["canonical_candidate_id"] == "candidate_001"
    assert artifact["source_refs"]["photo_ids"] == ["photo_001"]
    assert artifact["source_refs"]["note_item_ids"] == ["note_001"]
    assert artifact["source_refs"]["card_snapshot_ids"] == ["card_001"]
    assert artifact["proposed_writes"][0]["target_layer"] == "canonical structured state"
    assert artifact["proposed_writes"][0]["target_table"] == "mouse_master"
    assert artifact["proposed_writes"][0]["operation"] == "insert"
    assert artifact["proposed_writes"][0]["evidence_refs"] == ["note_001", "photo_001", "card_001"]
    assert artifact["blockers"] == []


def test_persist_validation_report_artifact_blocks_missing_trace(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    preview = {
        "candidate_id": "candidate_missing_trace",
        "candidate_status": "draft",
        "review_id": "review_002",
        "parse_id": "parse_002",
        "proposed_mice": [
            {
                "mouse_id": "mouse_MT319_parse_002",
                "display_id": "MT319",
                "source_note_item_id": "",
                "source_photo_id": "photo_002",
            }
        ],
        "duplicate_risks": [],
        "blockers": [],
    }

    report = app_main.build_canonical_apply_validation_report(
        preview,
        created_at="2026-05-09T12:05:00Z",
    )
    result = app_main.persist_validation_report_artifact(report)

    artifact_path = Path(result["artifact_path"])
    assert artifact_path.exists()
    assert artifact_path.parent == tmp_path / "mousedb_artifacts" / "validation_reports"

    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
    assert artifact["artifact_type"] == "validation_report"
    assert artifact["source_layer"] == "export or view"
    assert artifact["scope"] == "canonical_apply"
    assert artifact["status"] == "blocked"
    assert artifact["canonical_candidate_id"] == "candidate_missing_trace"
    assert artifact["source_refs"]["photo_ids"] == ["photo_002"]
    assert artifact["source_refs"]["parse_ids"] == ["parse_002"]
    assert any(
        check["check_key"] == "missing_source_trace" and check["status"] == "blocked"
        for check in artifact["checks"]
    )


def test_export_validation_report_blocks_open_focus_reviews(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    preview = {
        "export_type": "separation_preview",
        "blocked_review_items": 2,
        "latest_data_change_at": "2026-05-09T11:30:00Z",
        "expected_separation_filename": "2026-05-09 ApoM TgTg 분리 현황표.xlsx",
        "review_blockers": [
            {
                "review_id": "review_blocker_001",
                "severity": "High",
                "issue": "Duplicate active mouse",
            }
        ],
        "separation_rows": [
            {
                "source_photo_ids": "photo_010",
                "source_note_item_ids": "note_010, note_011",
            }
        ],
        "animal_sheet_rows": [],
    }

    report = app_main.build_export_validation_report(
        preview,
        export_type="separation_xlsx",
        query="ApoM TgTg",
        filename="2026-05-09 ApoM TgTg 분리 현황표.xlsx",
        created_at="2026-05-09T12:10:00Z",
    )
    result = app_main.persist_validation_report_artifact(report)

    artifact_path = Path(result["artifact_path"])
    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
    assert artifact["artifact_type"] == "validation_report"
    assert artifact["scope"] == "export"
    assert artifact["status"] == "blocked"
    assert artifact["state_watermark"] == "2026-05-09T11:30:00Z"
    assert artifact["source_refs"]["photo_ids"] == ["photo_010"]
    assert artifact["source_refs"]["note_item_ids"] == ["note_010", "note_011"]
    assert artifact["source_refs"]["review_ids"] == ["review_blocker_001"]
    assert any(
        check["check_key"] == "open_focus_review_blocker" and check["status"] == "blocked"
        for check in artifact["checks"]
    )


def test_export_validation_report_uses_uncompacted_preview_trace() -> None:
    preview = {
        "blocked_review_items": 0,
        "latest_data_change_at": "2026-05-09T11:40:00Z",
        "review_blockers": [],
        "preview_rows": [
            {
                "mouse_id": "mouse_010",
                "source_photo_id": "photo_010",
                "source_note_item_id": "note_010",
            },
            {
                "mouse_id": "mouse_011",
                "source_photo_id": "photo_011",
                "source_note_item_id": "note_011",
            },
            {
                "mouse_id": "mouse_012",
                "source_photo_id": "photo_012",
                "source_note_item_id": "note_012",
            },
            {
                "mouse_id": "mouse_013",
                "source_photo_id": "photo_013",
                "source_note_item_id": "note_013",
            },
        ],
        "separation_rows": [
            {
                "source_photo_ids": "photo_010, +3",
                "source_note_item_ids": "note_010, +3",
            }
        ],
        "animal_sheet_rows": [],
    }

    report = app_main.build_export_validation_report(
        preview,
        export_type="separation_xlsx",
        query="",
        filename="separation.xlsx",
        created_at="2026-05-09T12:10:00Z",
    )

    assert report["source_refs"]["photo_ids"] == ["photo_010", "photo_011", "photo_012", "photo_013"]
    assert report["source_refs"]["note_item_ids"] == ["note_010", "note_011", "note_012", "note_013"]
    assert report["source_refs"]["mouse_ids"] == ["mouse_010", "mouse_011", "mouse_012", "mouse_013"]
    assert "+3" not in report["checks"][1]["evidence_refs"]


def test_export_validation_report_warns_on_partially_untraced_rows() -> None:
    preview = {
        "blocked_review_items": 0,
        "latest_data_change_at": "2026-05-09T11:40:00Z",
        "review_blockers": [],
        "preview_rows": [
            {
                "mouse_id": "mouse_traced",
                "source_photo_id": "photo_traced",
                "source_note_item_id": "note_traced",
            },
            {
                "mouse_id": "mouse_untraced",
                "source_photo_id": "",
                "source_note_item_id": "",
            },
        ],
        "separation_rows": [
            {
                "source_photo_ids": "photo_traced",
                "source_note_item_ids": "note_traced",
            }
        ],
        "animal_sheet_rows": [],
    }

    report = app_main.build_export_validation_report(
        preview,
        export_type="separation_xlsx",
        query="",
        filename="separation.xlsx",
        created_at="2026-05-09T12:10:00Z",
    )

    missing_trace_check = next(
        check for check in report["checks"] if check["check_key"] == "missing_source_trace"
    )
    assert missing_trace_check["status"] == "warning"
    assert missing_trace_check["target_refs"] == ["mouse_untraced"]
    assert missing_trace_check["evidence_refs"] == ["photo_traced", "note_traced"]


def test_export_validation_report_does_not_treat_source_record_as_note_item() -> None:
    preview = {
        "blocked_review_items": 0,
        "latest_data_change_at": "2026-05-09T11:40:00Z",
        "review_blockers": [],
        "animal_sheet_rows": [
            {
                "mouse_id": "4p",
                "source": "source_litter_manual",
                "source_record_id": "source_litter_manual",
                "source_note_item_ids": "",
                "source_photo_ids": "",
            }
        ],
        "preview_rows": [],
        "separation_rows": [],
    }

    report = app_main.build_export_validation_report(
        preview,
        export_type="animal_sheet_xlsx",
        query="",
        filename="animal.xlsx",
        created_at="2026-05-09T12:10:00Z",
    )

    assert report["source_refs"]["note_item_ids"] == []
    assert report["source_refs"]["source_record_ids"] == ["source_litter_manual"]
    missing_trace_check = next(
        check for check in report["checks"] if check["check_key"] == "missing_source_trace"
    )
    assert missing_trace_check["status"] == "pass"
    assert missing_trace_check["evidence_refs"] == ["source_litter_manual"]


def test_export_validation_report_endpoint_uses_current_preview(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    monkeypatch.setattr(
        app_main,
        "export_preview",
        lambda: {
            "export_type": "separation_preview",
            "blocked_review_items": 0,
            "latest_data_change_at": "2026-05-09T11:45:00Z",
            "expected_separation_filename": "2026-05-09 selected strain 분리 현황표.xlsx",
            "review_blockers": [],
            "separation_rows": [
                {
                    "source_photo_ids": "photo_020",
                    "source_note_item_ids": "note_020",
                }
            ],
            "animal_sheet_rows": [],
        },
    )

    result = app_main.create_export_validation_report_artifact(
        export_type="separation_xlsx",
        query="",
    )

    artifact = result["artifact"]
    assert artifact["scope"] == "export"
    assert artifact["status"] == "pass"
    assert artifact["state_watermark"] == "2026-05-09T11:45:00Z"
    assert artifact["source_refs"]["photo_ids"] == ["photo_020"]


def test_persist_export_manifest_links_validation_report_and_sources(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    validation_report = {
        "report_id": "validation_report_export_separation_xlsx_ApoM",
        "artifact_path": str(tmp_path / "mousedb_artifacts" / "validation_reports" / "report.json"),
        "artifact": {
            "state_watermark": "2026-05-09T11:45:00Z",
            "source_refs": {
                "photo_ids": ["photo_030"],
                "note_item_ids": ["note_030"],
                "source_record_ids": ["source_030"],
                "review_ids": [],
                "mouse_ids": ["mouse_030"],
            },
            "status": "pass",
        },
    }

    result = app_main.persist_export_manifest_artifact(
        app_main.build_export_manifest(
            export_type="separation_xlsx",
            filename="2026-05-09 ApoM TgTg 분리 현황표.xlsx",
            query="ApoM TgTg",
            status="generated",
            row_count=1,
            blocked_review_count=0,
            validation_report=validation_report,
            created_at="2026-05-09T12:20:00Z",
        )
    )

    manifest_path = Path(result["artifact_path"])
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["artifact_type"] == "export_manifest"
    assert manifest["source_layer"] == "export or view"
    assert manifest["export_type"] == "separation_xlsx"
    assert manifest["filename"] == "2026-05-09 ApoM TgTg 분리 현황표.xlsx"
    assert manifest["status"] == "generated"
    assert manifest["validation_report_id"] == "validation_report_export_separation_xlsx_ApoM"
    assert manifest["validation_report_path"].endswith("report.json")
    assert manifest["state_watermark"] == "2026-05-09T11:45:00Z"
    assert manifest["source_refs"]["photo_ids"] == ["photo_030"]
    assert manifest["source_refs"]["note_item_ids"] == ["note_030"]
    assert manifest["source_refs"]["source_record_ids"] == ["source_030"]
    assert manifest["visual_qa"] == {
        "status": "manual_review_required",
        "automated_checks": [
            "workbook_structure",
            "trace_sheet_present",
            "source_refs_present",
        ],
        "manual_checks": [
            "lab_format_spacing",
            "printed_readability",
            "recipient_template_compatibility",
        ],
        "note": "Automated export checks do not replace manual lab-format workbook QA.",
    }


def test_log_workbook_export_preserves_manifest_provenance(tmp_path: Path) -> None:
    old_db_path = db.DB_PATH
    db.DB_PATH = tmp_path / "mouse_lims.sqlite"
    try:
        db.init_db()

        app_main.log_workbook_export(
            "separation_xlsx",
            "2026-05-09 ApoM TgTg 분리 현황표.xlsx",
            "ApoM TgTg",
            1,
            0,
            "generated",
            manifest_artifact_path="mousedb_artifacts/export_manifests/separation.json",
            validation_report_id="validation_report_export_separation_xlsx_ApoM",
            state_watermark="2026-05-09T11:45:00Z",
        )

        with db.connection() as conn:
            row = conn.execute(
                """
                SELECT export_type, filename, note
                FROM export_log
                ORDER BY exported_at DESC
                LIMIT 1
                """
            ).fetchone()

        assert row["export_type"] == "separation_xlsx"
        assert "manifest=mousedb_artifacts/export_manifests/separation.json" in row["note"]
        assert "validation_report=validation_report_export_separation_xlsx_ApoM" in row["note"]
        assert "state_watermark=2026-05-09T11:45:00Z" in row["note"]
    finally:
        db.DB_PATH = old_db_path


def test_export_log_api_exposes_structured_provenance(tmp_path: Path) -> None:
    old_db_path = db.DB_PATH
    old_artifact_root = app_main.ARTIFACT_ROOT
    db.DB_PATH = tmp_path / "mouse_lims.sqlite"
    artifact_root = tmp_path / "mousedb_artifacts"
    manifest_path = artifact_root / "export_manifests" / "animal_sheet.json"
    report_path = artifact_root / "validation_reports" / "animal_sheet_report.json"
    manifest_path.parent.mkdir(parents=True)
    report_path.parent.mkdir(parents=True)
    report_path.write_text(
        json.dumps({"artifact_type": "validation_report", "report_id": "validation_report_export_animal_sheet_xlsx_ApoM"}),
        encoding="utf-8",
    )
    manifest_path.write_text(
        json.dumps(
            {
                "artifact_type": "export_manifest",
                "validation_report_path": str(report_path),
            }
        ),
        encoding="utf-8",
    )
    try:
        db.init_db()

        app_main.log_workbook_export(
            "animal_sheet_xlsx",
            "2026-05-09 ApoM animalsheet.xlsx",
            "ApoM",
            2,
            0,
            "generated",
            manifest_artifact_path=str(manifest_path),
            validation_report_id="validation_report_export_animal_sheet_xlsx_ApoM",
            state_watermark="2026-05-09T12:05:00Z",
        )
        app_main.ARTIFACT_ROOT = artifact_root

        [row] = app_main.list_export_log()

        assert row["export_manifest_path"] == str(manifest_path)
        assert row["validation_report_id"] == "validation_report_export_animal_sheet_xlsx_ApoM"
        assert row["validation_report_path"] == str(report_path)
        assert row["state_watermark"] == "2026-05-09T12:05:00Z"
        assert row["provenance"] == {
            "export_manifest_path": str(manifest_path),
            "validation_report_id": "validation_report_export_animal_sheet_xlsx_ApoM",
            "validation_report_path": str(report_path),
            "state_watermark": "2026-05-09T12:05:00Z",
        }
    finally:
        app_main.ARTIFACT_ROOT = old_artifact_root
        db.DB_PATH = old_db_path


def test_artifact_preview_reads_json_under_artifact_root(tmp_path: Path, monkeypatch) -> None:
    artifact_root = tmp_path / "mousedb_artifacts"
    artifact_path = artifact_root / "export_manifests" / "manifest.json"
    artifact_path.parent.mkdir(parents=True)
    artifact_path.write_text(
        json.dumps(
            {
                "artifact_type": "export_manifest",
                "source_layer": "export or view",
                "manifest_id": "export_manifest_test",
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", artifact_root)

    preview = app_main.get_artifact_preview(str(artifact_path))

    assert preview["artifact_type"] == "export_manifest"
    assert preview["source_layer"] == "export or view"
    assert preview["relative_path"] == "export_manifests/manifest.json"
    assert preview["artifact"]["manifest_id"] == "export_manifest_test"


def test_artifact_preview_blocks_paths_outside_artifact_root(tmp_path: Path, monkeypatch) -> None:
    artifact_root = tmp_path / "mousedb_artifacts"
    outside_path = tmp_path / "outside.json"
    outside_path.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", artifact_root)

    with pytest.raises(HTTPException) as exc_info:
        app_main.get_artifact_preview(str(outside_path))

    assert exc_info.value.status_code == 400


def test_export_preview_reports_export_view_consistency_checks(tmp_path: Path) -> None:
    old_db_path = db.DB_PATH
    db.DB_PATH = tmp_path / "mouse_lims.sqlite"
    try:
        db.init_db()
        with db.connection() as conn:
            conn.execute(
                """
                INSERT INTO photo_log
                    (photo_id, original_filename, stored_path, uploaded_at, status, raw_source_kind)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    "photo_export_consistency",
                    "export-consistency-card.jpg",
                    "data/photos/test/export-consistency-card.jpg",
                    "2026-05-09T00:00:00Z",
                    "accepted",
                    "cage_card_photo",
                ),
            )
            conn.execute(
                """
                INSERT INTO parse_result
                    (parse_id, photo_id, source_name, raw_payload, parsed_at, status, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "parse_export_consistency",
                    "photo_export_consistency",
                    "manual_photo_transcription",
                    "{}",
                    "2026-05-09T00:00:01Z",
                    "accepted",
                    95,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO card_note_item_log
                    (note_item_id, photo_id, parse_id, card_snapshot_id, card_type,
                     line_number, raw_line_text, parsed_type, interpreted_status,
                     parsed_mouse_display_id, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "note_export_consistency",
                    "photo_export_consistency",
                    "parse_export_consistency",
                    "card_export_consistency",
                    "Separated",
                    1,
                    "MT401 R'",
                    "mouse_item",
                    "active",
                    "MT401",
                    95,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO mouse_master
                    (mouse_id, display_id, raw_strain_text, sex, dob_raw,
                     source_note_item_id, current_card_snapshot_id, status,
                     source_photo_id, last_verified_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mouse_export_consistency",
                    "MT401",
                    "ApoM Tg/Tg",
                    "female",
                    "2026-04-01",
                    "note_export_consistency",
                    "card_export_consistency",
                    "active",
                    "photo_export_consistency",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                ),
            )

        preview = app_main.export_preview()

        assert preview["source_layer"] == "export or view"
        assert preview["row_state_policy"] == {
            "source_layer": "export or view",
            "source_state_layer": "canonical structured state",
            "states": ["ready", "blocked_by_review", "stale_after_correction"],
            "editable": False,
        }
        assert preview["preview_rows"][0]["row_state"] == "ready"
        assert preview["preview_rows"][0]["row_state_reason"] == "Canonical row is ready for Excel export."
        assert preview["separation_rows"][0]["row_state"] == "ready"
        assert preview["animal_sheet_rows"] == []
        assert preview["export_consistency"] == {
            "source_layer": "export or view",
            "source_state_layer": "canonical structured state",
            "preview_row_count": 1,
            "separation_row_count": 1,
            "animal_sheet_row_count": 0,
            "preview_rows_have_trace": True,
            "separation_rows_have_trace": True,
            "animal_sheet_rows_have_trace": "not_applicable",
            "excel_export_is_view": True,
            "preview_rows_have_row_state": True,
            "separation_rows_have_row_state": True,
            "animal_sheet_rows_have_row_state": "not_applicable",
        }
    finally:
        db.DB_PATH = old_db_path


def test_export_row_state_uses_staleness_contract_and_blockers_take_priority() -> None:
    assert app_main.export_row_state(0, {"export_stale": True, "latest_generated_export_at": ""}) == {
        "row_state": "ready",
        "row_state_reason": "Canonical row is ready for Excel export.",
    }
    assert app_main.export_row_state(
        0,
        {
            "export_stale": True,
            "latest_generated_export_at": "2026-05-08T00:00:00Z",
        },
    ) == {
        "row_state": "stale_after_correction",
        "row_state_reason": "Accepted state changed after the latest export.",
    }
    assert app_main.export_row_state(
        2,
        {
            "export_stale": True,
            "latest_generated_export_at": "2026-05-08T00:00:00Z",
        },
    ) == {
        "row_state": "blocked_by_review",
        "row_state_reason": "Focus Review blockers remain before Excel export.",
    }


def test_export_preview_ui_consumes_backend_row_state_contract() -> None:
    static_html = Path("static/index.html").read_text(encoding="utf-8")

    start = static_html.index("function exportPreviewRowStateChips")
    row_state_function = static_html[
        start : static_html.index("function exportPreviewRow", start + 1)
    ]

    assert "item?.row_state" in row_state_function
    assert "row_state_reason" in row_state_function
    assert "stale_after_correction" in row_state_function


def test_static_export_center_prioritizes_final_gate_and_download_hierarchy() -> None:
    static_html = Path("static/index.html").read_text(encoding="utf-8")

    export_center_index = static_html.index("<h2>Export Center</h2>")
    search_export_index = static_html.index("<h2>Search & CSV Export</h2>")
    assert export_center_index < search_export_index

    start = static_html.rindex('<section class="full" data-view="exports">', 0, export_center_index)
    end = static_html.index("</section>", export_center_index)
    export_center = static_html[start:end]

    assert 'class="export-gate-banner"' in export_center
    assert "Preview / search" in export_center
    assert "Worklists" in export_center
    assert "Final lab files" in export_center
    assert "Preview-only" in export_center
    assert "Worklist" in export_center
    assert "Final lab file" in export_center
    assert export_center.index('id="exportReadinessCard"') < export_center.index('id="exportMouseCsvButton"')
    assert export_center.index('id="exportReadinessCard"') < export_center.index('id="exportGenotypingCsvButton"')
    assert export_center.index("Final lab files") < export_center.index('id="exportReadyMouseCsvButton"')


def test_static_export_final_lab_files_copy_tracks_gate_state() -> None:
    static_html = Path("static/index.html").read_text(encoding="utf-8")

    assert 'id="exportFinalLabFilesGroup"' in static_html
    assert 'id="exportFinalLabFilesCopy"' in static_html
    start = static_html.index("function setFinalExportActionState")
    setter = static_html[start : static_html.index("function renderExportReadiness", start)]

    assert "exportFinalLabFilesCopy" in setter
    assert "Final lab file downloads are ready" in setter
    assert "Final lab file downloads stay disabled" in setter
    assert "exportFinalLabFilesGroup" in setter
    assert 'classList.toggle("ready", ready)' in setter
    assert 'classList.toggle("blocked", !ready)' in setter


def test_separation_xlsx_renders_trace_sheet_with_row_state_and_source_refs(
    tmp_path: Path,
    monkeypatch,
) -> None:
    old_db_path = db.DB_PATH
    db.DB_PATH = tmp_path / "mouse_lims.sqlite"
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    try:
        db.init_db()
        with db.connection() as conn:
            conn.execute(
                """
                INSERT INTO photo_log
                    (photo_id, original_filename, stored_path, uploaded_at, status, raw_source_kind)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    "photo_xlsx_render",
                    "xlsx-render-card.jpg",
                    "data/photos/test/xlsx-render-card.jpg",
                    "2026-05-09T00:00:00Z",
                    "accepted",
                    "cage_card_photo",
                ),
            )
            conn.execute(
                """
                INSERT INTO parse_result
                    (parse_id, photo_id, source_name, raw_payload, parsed_at, status, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "parse_xlsx_render",
                    "photo_xlsx_render",
                    "manual_photo_transcription",
                    "{}",
                    "2026-05-09T00:00:01Z",
                    "accepted",
                    96,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO card_note_item_log
                    (note_item_id, photo_id, parse_id, card_snapshot_id, card_type,
                     line_number, raw_line_text, parsed_type, interpreted_status,
                     parsed_mouse_display_id, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "note_xlsx_render",
                    "photo_xlsx_render",
                    "parse_xlsx_render",
                    "card_xlsx_render",
                    "Separated",
                    1,
                    "MT777 R'",
                    "mouse_item",
                    "active",
                    "MT777",
                    96,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO mouse_master
                    (mouse_id, display_id, raw_strain_text, sex, dob_raw,
                     source_note_item_id, current_card_snapshot_id, status,
                     source_photo_id, last_verified_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mouse_xlsx_render",
                    "MT777",
                    "ApoM Tg/Tg",
                    "female",
                    "2026-04-01",
                    "note_xlsx_render",
                    "card_xlsx_render",
                    "active",
                    "photo_xlsx_render",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                ),
            )

        response = TestClient(app_main.app).get("/api/exports/separation.xlsx")

        assert response.status_code == 200
        assert response.content[:4] == b"PK\x03\x04"
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        assert "Export_Trace" in workbook.sheetnames
        trace = workbook["Export_Trace"]
        headers = [trace.cell(1, column).value for column in range(1, 12)]
        assert headers == [
            "Row",
            "Source note",
            "Source record",
            "Boundary",
            "Export note",
            "Source photo",
            "Card snapshot",
            "Raw note line",
            "Uncertainty",
            "Row state",
            "Row state reason",
        ]
        assert trace.cell(2, 2).value == "note_xlsx_render"
        assert trace.cell(2, 4).value == "export or view"
        assert trace.cell(2, 6).value == "photo_xlsx_render"
        assert trace.cell(2, 8).value == "MT777 R'"
        assert trace.cell(2, 10).value == "ready"
        assert trace.cell(2, 11).value == "Canonical row is ready for Excel export."
    finally:
        db.DB_PATH = old_db_path


def test_animal_sheet_xlsx_renders_trace_sheet_with_row_state_and_source_refs(
    tmp_path: Path,
    monkeypatch,
) -> None:
    old_db_path = db.DB_PATH
    db.DB_PATH = tmp_path / "mouse_lims.sqlite"
    monkeypatch.setattr(app_main, "ARTIFACT_ROOT", tmp_path / "mousedb_artifacts")
    try:
        db.init_db()
        with db.connection() as conn:
            conn.execute(
                """
                INSERT INTO photo_log
                    (photo_id, original_filename, stored_path, uploaded_at, status, raw_source_kind)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    "photo_animal_render",
                    "animal-render-card.jpg",
                    "data/photos/test/animal-render-card.jpg",
                    "2026-05-09T00:00:00Z",
                    "accepted",
                    "cage_card_photo",
                ),
            )
            conn.execute(
                """
                INSERT INTO parse_result
                    (parse_id, photo_id, source_name, raw_payload, parsed_at, status, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "parse_animal_render",
                    "photo_animal_render",
                    "manual_photo_transcription",
                    "{}",
                    "2026-05-09T00:00:01Z",
                    "accepted",
                    96,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO card_note_item_log
                    (note_item_id, photo_id, parse_id, card_snapshot_id, card_type,
                     line_number, raw_line_text, parsed_type, interpreted_status,
                     parsed_mouse_display_id, confidence, needs_review)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "note_animal_render",
                    "photo_animal_render",
                    "parse_animal_render",
                    "card_animal_render",
                    "mating",
                    1,
                    "MT888 R'",
                    "mouse_item",
                    "mating",
                    "MT888",
                    96,
                    0,
                ),
            )
            conn.execute(
                """
                INSERT INTO mouse_master
                    (mouse_id, display_id, raw_strain_text, sex, dob_raw,
                     source_note_item_id, current_card_snapshot_id, status,
                     source_photo_id, last_verified_at, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mouse_animal_render",
                    "MT888",
                    "ApoM Tg/Tg",
                    "male",
                    "2026-03-01",
                    "note_animal_render",
                    "card_animal_render",
                    "mating",
                    "photo_animal_render",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                    "2026-05-09T00:00:02Z",
                ),
            )
            conn.execute(
                """
                INSERT INTO mating_registry
                    (mating_id, mating_label, strain_goal, expected_genotype,
                     start_date, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mating_animal_render",
                    "Mating A",
                    "ApoM Tg/Tg",
                    "Tg/Tg",
                    "2026-05-01",
                    "active",
                    "2026-05-09T00:00:03Z",
                    "2026-05-09T00:00:03Z",
                ),
            )
            conn.execute(
                """
                INSERT INTO mating_mouse
                    (mating_mouse_id, mating_id, mouse_id, role,
                     joined_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    "mating_mouse_animal_render",
                    "mating_animal_render",
                    "mouse_animal_render",
                    "male",
                    "2026-05-01",
                    "2026-05-09T00:00:04Z",
                ),
            )

        response = TestClient(app_main.app).get("/api/exports/animal-sheet.xlsx")

        assert response.status_code == 200
        assert response.content[:4] == b"PK\x03\x04"
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        assert "animal sheet" in workbook.sheetnames
        assert "Export_Trace" in workbook.sheetnames
        animal_sheet = workbook["animal sheet"]
        assert [animal_sheet.cell(1, column).value for column in range(1, 11)] == [
            "Cage No.",
            "Strain",
            "Sex",
            "I.D",
            "genotype",
            "DOB",
            "Mating date",
            "Pubs",
            "Status",
            "Source",
        ]
        assert animal_sheet.cell(2, 4).value == "MT888"
        trace = workbook["Export_Trace"]
        assert [trace.cell(1, column).value for column in range(1, 12)] == [
            "Row",
            "Source note",
            "Source record",
            "Boundary",
            "Export note",
            "Source photo",
            "Card snapshot",
            "Raw note line",
            "Uncertainty",
            "Row state",
            "Row state reason",
        ]
        assert trace.cell(2, 2).value == "note_animal_render"
        assert trace.cell(2, 4).value == "export or view"
        assert trace.cell(2, 6).value == "photo_animal_render"
        assert trace.cell(2, 8).value == "MT888 R'"
        assert trace.cell(2, 10).value == "ready"
    finally:
        db.DB_PATH = old_db_path
