from __future__ import annotations

import json
import io
import sqlite3
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

try:
    from fastapi.testclient import TestClient
except (ModuleNotFoundError, RuntimeError):
    TestClient = None

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:
    Workbook = None
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
CLI_MAIN = ROOT / "mousedb" / "__main__.py"
CLI_PYTHON = ROOT / ".venv" / "Scripts" / "python.exe"


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def run_cli(data_dir: Path, *args: str, expect_code: int = 0) -> subprocess.CompletedProcess[str]:
    python_executable = str(CLI_PYTHON) if CLI_PYTHON.exists() else sys.executable
    result = subprocess.run(
        [python_executable, "-m", "mousedb", "--db", str(data_dir / "mousedb.sqlite"), *args],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )
    assert_true(
        result.returncode == expect_code,
        f"CLI {' '.join(args)} returned {result.returncode}: stdout={result.stdout!r} stderr={result.stderr!r}",
    )
    return result


def main() -> None:
    for path in [
        ROOT / "app" / "main.py",
        ROOT / "app" / "db.py",
        ROOT / "app" / "storage.py",
        CLI_MAIN,
        ROOT / "static" / "index.html",
        ROOT / "requirements.txt",
        ROOT / "start.bat",
    ]:
        assert_true(path.exists(), f"Missing required local app file: {path}")

    fixture = json.loads((ROOT / "fixtures" / "sample_parse_results.json").read_text(encoding="utf-8"))
    assert_true(fixture.get("layer") == "parsed or intermediate result", "Fixture must stay non-canonical.")
    assert_true(len(fixture.get("records", [])) >= 3, "Fixture should contain parse records.")
    assert_true(Workbook is not None and load_workbook is not None, "openpyxl is required for workbook parsing and validation.")

    with tempfile.TemporaryDirectory() as source_dir:
        animal_path = Path(source_dir) / "legacy_animal.xlsx"
        animal_workbook = Workbook()
        animal_sheet = animal_workbook.active
        animal_sheet.title = "animal sheet"
        animal_sheet.append(["Cage No.", "Strain", "Sex", "I.D", "genotype", "DOB", "Mating date", "Pubs"])
        animal_sheet.append(["1", "ApoM Tg/Tg", "M", "MT321", "Tg/Tg", "2026-01-01", "2026-05-01", ""])
        animal_sheet.append(["", "", "F1", "9p", "pre_weaning", "2026-05-02", "", "2026-05-02 9p"])
        animal_workbook.save(animal_path)
        python_executable = str(CLI_PYTHON) if CLI_PYTHON.exists() else sys.executable
        animal_result = subprocess.run(
            [python_executable, str(ROOT / "scripts" / "parse_legacy_workbooks.py"), str(animal_path), "--kind", "animal"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        assert_true(animal_result.returncode == 0, f"Legacy animal parser failed: {animal_result.stderr}")
        animal_payload = json.loads(animal_result.stdout)
        assert_true(animal_payload["layer"] == "parsed or intermediate result", "Legacy workbook parser should stay non-canonical.")
        assert_true(animal_payload["source_layer"] == "export or view", "Legacy workbook source should be classified as export/view.")
        assert_true(animal_payload["rows"][0]["row_type"] == "parent_or_mouse_snapshot", "Animal parser should classify M rows as mouse snapshots.")
        assert_true(animal_payload["rows"][1]["row_type"] == "litter_or_offspring_snapshot", "Animal parser should classify F1 rows as litter snapshots.")
        assert_true(animal_payload["rows"][0]["source_cells"]["display_id"] == "D2", "Animal parser should preserve cell traceability.")

        separation_path = Path(source_dir) / "legacy_separation.xlsx"
        separation_workbook = Workbook()
        separation_sheet = separation_workbook.active
        separation_sheet.title = "separation"
        separation_sheet.append(["Strain", "Genotype", "total", "DOB", "WT", "Tg", "Sampling point"])
        separation_sheet.append(["ApoM Tg/Tg", "Tg/Tg", "M 3p", "2026-01-01", "", "3", "tail"])
        separation_workbook.save(separation_path)
        separation_result = subprocess.run(
            [python_executable, str(ROOT / "scripts" / "parse_legacy_workbooks.py"), str(separation_path), "--kind", "separation"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        )
        assert_true(separation_result.returncode == 0, f"Legacy separation parser failed: {separation_result.stderr}")
        separation_payload = json.loads(separation_result.stdout)
        assert_true(separation_payload["workbook_kind"] == "legacy_separation_status", "Separation parser should label workbook kind.")
        assert_true(separation_payload["rows"][0]["sex_candidate"] == "male", "Separation parser should infer ASCII M sex labels.")
        assert_true(separation_payload["rows"][0]["count_candidate"] == 3, "Separation parser should infer pup counts.")

    sys.path.insert(0, str(ROOT))
    from app import db
    from scripts.parse_legacy_workbooks import parse_workbook

    with tempfile.TemporaryDirectory() as workbook_dir:
        workbook_root = Path(workbook_dir)
        animal_path = workbook_root / "legacy_animal.xlsx"
        animal_wb = Workbook()
        animal_ws = animal_wb.active
        animal_ws.title = "ApoM TgTg"
        animal_ws.append(["Cage No.", "Strain", "Sex", "I.D", "genotype", "DOB", "Mating date", "Pubs"])
        animal_ws.append(["1", "ApoM Tg/Tg", "\u2642", "MT 318 R'", "Tg", "25.10.20-28", "26.01.30", ""])
        animal_ws.append(["", "", "F1", "7p", "separated", "26.02.04", "", ""])
        animal_wb.save(animal_path)
        animal_payload = parse_workbook(animal_path, kind="animal")
        assert_true(animal_payload["layer"] == "parsed or intermediate result", "Legacy animal parser must stay non-canonical.")
        assert_true(animal_payload["source_layer"] == "export or view", "Legacy animal workbook must be classified as a view.")
        assert_true(len(animal_payload["rows"]) == 2, "Legacy animal parser row count is wrong.")
        assert_true(animal_payload["rows"][0]["source_cells"]["display_id"] == "D2", "Legacy animal parser must preserve cell traceability.")

        separation_path = workbook_root / "legacy_separation.xlsx"
        separation_wb = Workbook()
        separation_ws = separation_wb.active
        separation_ws.title = "ApoM TgTg"
        separation_ws.append(["Strain", "Genotype", "total", "DOB", "Genotype", "", "", "Sampling point"])
        separation_ws.append(["", "", "", "", "WT", "Tg", "", "10mths note"])
        separation_ws.append(["Apom Tg/Tg", "Apom Tg/Tg", "\u2642 2p", "25.05.07", "", "2", "", ""])
        separation_ws.append(["", "Apom Tg/Tg", "\u2640 6p", "26.02.18-24", "", "6", "", ""])
        separation_wb.save(separation_path)
        separation_payload = parse_workbook(separation_path, kind="separation")
        assert_true(separation_payload["source_layer"] == "export or view", "Legacy separation workbook must be classified as a view.")
        assert_true(len(separation_payload["rows"]) == 2, "Legacy separation parser should skip subheader-only rows.")
        assert_true(separation_payload["rows"][0]["count_candidate"] == 2, "Legacy separation parser should extract total counts as candidates.")
        assert_true(separation_payload["rows"][1]["sex_candidate"] == "female", "Legacy separation parser should extract sex as a candidate.")

    app = None
    if TestClient is not None:
        from app.main import app

    with tempfile.TemporaryDirectory() as old_schema_dir:
        db.DATA_DIR = Path(old_schema_dir)
        db.DB_PATH = Path(old_schema_dir) / "mouse_lims.sqlite"
        legacy_conn = sqlite3.connect(db.DB_PATH)
        try:
            legacy_conn.executescript(
                """
                CREATE TABLE mouse_master (
                    mouse_id TEXT PRIMARY KEY,
                    display_id TEXT NOT NULL
                );
                CREATE TABLE card_note_item_log (
                    note_item_id TEXT PRIMARY KEY,
                    raw_line_text TEXT NOT NULL
                );
                CREATE TABLE genotyping_record (
                    genotyping_id TEXT PRIMARY KEY,
                    mouse_id TEXT,
                    sample_id TEXT,
                    sample_date TEXT,
                    result_date TEXT,
                    created_at TEXT
                );
                """
            )
            legacy_conn.commit()
        finally:
            legacy_conn.close()
        db.init_db()
        migrated_conn = sqlite3.connect(db.DB_PATH)
        try:
            mouse_columns = {
                row[1]
                for row in migrated_conn.execute("PRAGMA table_info(mouse_master)").fetchall()
            }
            note_columns = {
                row[1]
                for row in migrated_conn.execute("PRAGMA table_info(card_note_item_log)").fetchall()
            }
            genotype_columns = {
                row[1]
                for row in migrated_conn.execute("PRAGMA table_info(genotyping_record)").fetchall()
            }
        finally:
            migrated_conn.close()
        assert_true(
            {
                "target_match_status",
                "use_category",
                "next_action",
                "sample_id",
                "father_id",
                "mother_id",
                "litter_id",
                "source_record_id",
            }.issubset(mouse_columns),
            "Existing mouse_master tables should migrate to the genotyping, lineage, and traceability schema.",
        )
        assert_true(
            {"parsed_ear_label_code", "strike_status", "needs_review"}.issubset(note_columns),
            "Existing card_note_item_log tables should migrate to the note parsing schema.",
        )
        assert_true(
            {"target_name", "normalized_result", "result_status", "updated_at"}.issubset(genotype_columns),
            "Existing genotyping_record tables should migrate to the result tracking schema.",
        )

    if CLI_PYTHON.exists():
        with tempfile.TemporaryDirectory() as cli_dir:
            cli_data_dir = Path(cli_dir)
            initialized = json.loads(run_cli(cli_data_dir, "init", "--json").stdout)
            assert_true(initialized["initialized"] is True, "MouseDB CLI init should initialize the local database.")
            strain = json.loads(
                run_cli(
                    cli_data_dir,
                    "strain",
                    "add",
                    "--name",
                    "ApoM Tg/Tg",
                    "--source",
                    "manual",
                    "--json",
                ).stdout
            )
            assert_true(strain["strain_id"].startswith("STR-"), "MouseDB CLI strain add should create an external strain ID.")
            cage = json.loads(
                run_cli(
                    cli_data_dir,
                    "cage",
                    "add",
                    "--label",
                    "C-014",
                    "--type",
                    "holding",
                    "--json",
                ).stdout
            )
            assert_true(cage["cage_id"] == "C-014", "MouseDB CLI cage add should normalize cage IDs.")
            mouse = json.loads(
                run_cli(
                    cli_data_dir,
                    "mouse",
                    "add",
                    "--display-id",
                    "MT321",
                    "--strain",
                    strain["strain_id"],
                    "--sex",
                    "F",
                    "--dob",
                    "2025-10-20",
                    "--cage",
                    cage["cage_id"],
                    "--json",
                ).stdout
            )
            assert_true(mouse["display_id"] == "MT321", "MouseDB CLI mouse add should preserve display ID.")
            mate = json.loads(
                run_cli(
                    cli_data_dir,
                    "mouse",
                    "add",
                    "--display-id",
                    "MT322",
                    "--strain",
                    strain["strain_id"],
                    "--sex",
                    "M",
                    "--dob",
                    "2025-10-20",
                    "--cage",
                    cage["cage_id"],
                    "--json",
                ).stdout
            )
            assert_true(mate["mouse_id"] != mouse["mouse_id"], "MouseDB CLI should create distinct mouse IDs.")
            genotype = json.loads(
                run_cli(
                    cli_data_dir,
                    "genotype",
                    "record",
                    "--mouse",
                    mouse["mouse_id"],
                    "--result",
                    "Tg/Tg",
                    "--sample-id",
                    "S-MT321",
                    "--json",
                ).stdout
            )
            assert_true(genotype["result"] == "Tg/Tg", "MouseDB CLI genotype record should preserve result text.")
            mating = json.loads(
                run_cli(
                    cli_data_dir,
                    "mating",
                    "create",
                    "--male",
                    mate["mouse_id"],
                    "--female",
                    mouse["mouse_id"],
                    "--goal",
                    strain["strain_name"],
                    "--expected-genotype",
                    "Tg/Tg",
                    "--json",
                ).stdout
            )
            assert_true(mating["status"] == "active", "MouseDB CLI mating create should create an active mating.")
            litter = json.loads(
                run_cli(
                    cli_data_dir,
                    "litter",
                    "create",
                    "--mating",
                    mating["mating_id"],
                    "--number-born",
                    "6",
                    "--json",
                ).stdout
            )
            assert_true(litter["number_born"] == 6, "MouseDB CLI litter create should preserve litter counts.")
            summary = json.loads(run_cli(cli_data_dir, "colony", "summary", "--json").stdout)
            assert_true(summary["total_alive_mice"] >= 2, "MouseDB CLI colony summary should include live mouse totals.")
            assert_true(summary["active_matings"] == 1, "MouseDB CLI colony summary should include active mating counts.")

    with tempfile.TemporaryDirectory() as temp_dir:
        db.DATA_DIR = Path(temp_dir)
        db.DB_PATH = Path(temp_dir) / "mouse_lims.sqlite"
        db.init_db()
        conn = sqlite3.connect(db.DB_PATH)
        try:
            tables = {
                row[0]
                for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }
        finally:
            conn.close()
        assert_true(
            {
                "photo_log",
                "parse_result",
                "review_queue",
                "action_log",
                "source_record",
                "strain_registry",
                "correction_log",
                "export_log",
                "mouse_event",
                "genotyping_record",
                "strain_target_genotype",
                "cage_registry",
                "mouse_cage_assignment",
                "mating_registry",
                "mating_mouse",
                "litter_registry",
                "my_assigned_strain",
                "distribution_import",
                "distribution_assignment_row",
                "ear_label_master",
                "ear_label_alias",
                "mouse_master",
                "card_note_item_log",
            }.issubset(tables),
            "Local SQLite schema is incomplete.",
        )
        conn = sqlite3.connect(db.DB_PATH)
        try:
            master_rows = dict(
                conn.execute("SELECT ear_label_code, display_text FROM ear_label_master").fetchall()
            )
            ambiguous_alias = conn.execute(
                """
                SELECT confirmed
                FROM ear_label_alias
                WHERE raw_text = ? AND ear_label_code = ?
                """,
                ("R0", "R_CIRCLE"),
            ).fetchone()
            conn.execute(
                """
                INSERT INTO card_note_item_log
                    (note_item_id, line_number, raw_line_text, parsed_type, strike_status,
                     interpreted_status, parsed_mouse_display_id, parsed_ear_label_raw,
                     parsed_ear_label_code, confidence)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "note_test_319",
                    1,
                    "319 L'",
                    "mouse_item",
                    "none",
                    "active",
                    "319",
                    "L'",
                    "L_PRIME",
                    0.98,
                ),
            )
            conn.execute(
                """
                INSERT INTO mouse_master
                    (mouse_id, display_id, raw_strain_text, dob_raw, dob_start, dob_end,
                     ear_label_raw, ear_label_code, source_note_item_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "mouse_test_319_a",
                    "319",
                    "ApoM Tg/Tg",
                    "25.10.20-28",
                    "2025-10-20",
                    "2025-10-28",
                    "L'",
                    "L_PRIME",
                    "note_test_319",
                ),
            )
            conn.execute(
                """
                INSERT INTO mouse_master
                    (mouse_id, display_id, raw_strain_text, dob_start, ear_label_code)
                VALUES (?, ?, ?, ?, ?)
                """,
                ("mouse_test_319_b", "319", "Different strain candidate", "2026-01-01", "R_PRIME"),
            )
            mouse_defaults = conn.execute(
                """
                SELECT genotyping_status, next_action, status
                FROM mouse_master
                WHERE mouse_id = ?
                """,
                ("mouse_test_319_a",),
            ).fetchone()
            same_display_count = conn.execute(
                "SELECT COUNT(*) FROM mouse_master WHERE display_id = ?",
                ("319",),
            ).fetchone()[0]
            conn.commit()
        finally:
            conn.close()
        assert_true(master_rows.get("R_CIRCLE") == "R\u00b0", "R_CIRCLE must use degree-sign display text.")
        assert_true(master_rows.get("L_CIRCLE") == "L\u00b0", "L_CIRCLE must use degree-sign display text.")
        assert_true(ambiguous_alias is not None, "Ambiguous R0 alias should be seeded for review.")
        assert_true(ambiguous_alias[0] == 0, "Ambiguous R0 alias must not be auto-confirmed.")
        assert_true(
            tuple(mouse_defaults) == ("not_sampled", "sample_needed", "active"),
            "Mouse workflow defaults are wrong.",
        )
        assert_true(same_display_count == 2, "Mouse display IDs must remain non-unique identity candidates.")

        if TestClient is None:
            with db.connection() as conn:
                conn.execute(
                    """
                    INSERT INTO my_assigned_strain
                        (assigned_strain_id, display_name, aliases_json, source_type, assigned_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    ("assigned_strain_test", "ApoM Tg/Tg", '["ApoMtg/tg"]', "manual", "test"),
                )
                count = conn.execute("SELECT COUNT(*) AS count FROM my_assigned_strain").fetchone()["count"]
            assert_true(count == 1, "Assigned strain scope table did not accept a row.")
        else:
            with TestClient(app) as client:
                index_html = client.get("/").text
                assert_true("Colony Records" in index_html, "Local UI should expose mouse records.")
                assert_true("Parsed Note Evidence" in index_html, "Local UI should expose parsed note evidence.")
                assert_true("Strain Registry" in index_html, "Local UI should expose strain registry.")
                assert_true("Source Evidence" in index_html, "Local UI should expose source evidence.")
                assert_true("Mouse Events" in index_html, "Local UI should expose mouse events.")
                assert_true("Correction Log" in index_html, "Local UI should expose correction history.")
                assert_true("Mouse Audit Trace" in index_html, "Local UI should expose per-mouse audit trace.")
                assert_true("audit-trail" in index_html, "Local UI should call the per-mouse audit trail API.")
                assert_true("Search & CSV Export" in index_html, "Local UI should expose search and CSV export.")
                assert_true("Download Genotyping Worklist" in index_html, "Local UI should expose genotyping worklist export.")
                assert_true("Download Ready CSV" in index_html, "Local UI should expose gated final CSV export.")
                assert_true("Cage View" in index_html, "Local UI should expose cage management.")
                assert_true("Breeding / Litter View" in index_html, "Local UI should expose mating and litter management.")
                assert_true("Create Offspring" in index_html, "Local UI should expose litter offspring generation.")
                assert_true("Complete Weaning" in index_html, "Local UI should expose litter weaning completion.")
                assert_true("Request Genotyping" in index_html, "Local UI should expose genotyping request workflow.")
                assert_true("Target genotype" in index_html, "Local UI should expose configurable target genotype rules.")
                assert_true("genotypingDashboard" in index_html, "Local UI should expose genotyping dashboard cards.")
                assert_true("exportRows" in index_html, "Local UI should expose export preview rows.")
                assert_true("exportFilenames" in index_html, "Local UI should expose expected workbook filenames.")
                assert_true("exportBlockerRows" in index_html, "Local UI should expose export blockers.")
                assert_true("exportLogRows" in index_html, "Local UI should expose export history.")
                assert_true("Review Queue" in index_html and "Evidence" in index_html, "Local UI should show review evidence context.")
                assert_true("reviewStatusFilter" in index_html, "Local UI should expose review status filtering.")
                assert_true("reviewSeverityFilter" in index_html, "Local UI should expose review severity filtering.")
                assert_true("reviewEvidenceFilter" in index_html, "Local UI should expose review evidence filtering.")
                assert_true("review-resolution-note" in index_html, "Local UI should resolve reviews inline instead of using prompt-only workflow.")
                assert_true("Mouse Audit Trace" in index_html, "Local UI should expose mouse audit trace view.")
                assert_true("auditTraceRows" in index_html, "Local UI should render audit trace rows.")
                assert_true("Deactivate" in index_html, "Local UI should expose assigned strain deactivation.")
                assert_true("Distribution Assignment Import" in index_html, "Local UI should expose distribution import.")
                assert_true("Legacy Workbook Import" in index_html, "Local UI should expose legacy workbook import.")
                assert_true("legacyWorkbookKind" in index_html, "Local UI should expose legacy workbook kind selection.")
                assert_true("legacyWorkbookRows" in index_html, "Local UI should render legacy workbook rows.")
                assert_true("Colony Dashboard" in index_html, "Local UI should expose the colony visualization dashboard.")
                assert_true("Mouse Detail" in index_html, "Local UI should expose the mouse detail visualization.")
                assert_true("Mouse Audit Trace" in index_html, "Local UI should expose mouse audit trace.")
                assert_true("auditTraceRows" in index_html, "Local UI should render audit trace rows.")
                assert_true("Strain Detail" in index_html, "Local UI should expose the strain detail visualization.")
                assert_true("Evidence & Review Readiness" in index_html, "Local UI should expose visualization evidence readiness.")
                assert_true("renderVisualizations" in index_html, "Local UI should render visualizations from API data.")
                assert_true("vizHeatmapHead" in index_html, "Genotype heatmap should be driven by rendered data labels.")
                assert_true("vizQualityRows" in index_html, "Visualization readiness should render data quality rows.")
                assert_true("demo-note-1" not in index_html, "Mouse detail visualization should not use demo source evidence.")
                assert_true("selectedStrainMice.length || aliveMice" not in index_html, "Strain detail active mice should not fall back to colony-wide counts.")
                assert_true("EXP-2026-041" not in index_html, "Strain visualization should not hard-code experiment IDs.")
                assert_true("/[^a-z0-9가-힣]/g" in index_html, "Local UI strain matching key should preserve Korean strain text.")
                assert_true(client.get("/api/assigned-strains").json() == [], "Assigned strain scope should start empty.")
                assert_true(client.get("/api/source-records").json() == [], "Source evidence should start empty.")
                assert_true(client.get("/api/strains").json() == [], "Strain registry should start empty.")
                strain = client.post(
                    "/api/strains",
                    json={
                        "strain_name": "PV-Cre",
                        "gene": "Pvalb",
                        "allele": "Pvalb-IRES-Cre",
                        "background": "C57BL/6J",
                        "source": "manual",
                        "status": "active",
                    },
                )
                assert_true(strain.status_code == 200, "Could not create strain registry entry.")
                strain_payload = strain.json()
                assert_true(strain_payload["source_record_id"], "Strain entry should create source evidence.")
                strains = client.get("/api/strains").json()
                assert_true(strains[0]["strain_name"] == "PV-Cre", "Strain registry did not persist entry.")
                source_records = client.get("/api/source-records").json()
                assert_true(
                    any(item["source_type"] == "manual_entry" for item in source_records),
                    "Manual strain creation should leave source evidence.",
                )
                distribution = client.post(
                    "/api/distribution-imports",
                    json={
                        "layer": "parsed or intermediate result",
                        "source_file_name": "distribution_test.xlsx",
                        "sheet_name": "Mating",
                        "rows": [
                            {
                                "institution_or_group": "Vet Med",
                                "responsible_person_raw": "Jang S.",
                                "mating_type_raw": "ApoMtg/tg",
                                "cage_count_raw": "6",
                                "source_row_number": 35,
                                "source_cells": {"mating_type": "B35"},
                                "review_status": "candidate",
                            }
                        ],
                    },
                )
                assert_true(distribution.status_code == 200, "Could not store distribution import JSON.")
                distribution_payload = distribution.json()
                assert_true(distribution_payload["stored_rows"] == 1, "Distribution import row count is wrong.")
                assert_true(distribution_payload["source_record_id"], "Distribution import should create source evidence.")
                distribution_imports = client.get("/api/distribution-imports").json()
                stored_distribution = next(
                    (
                        item
                        for item in distribution_imports
                        if item["distribution_import_id"] == distribution_payload["distribution_import_id"]
                    ),
                    None,
                )
                assert_true(stored_distribution is not None, "Distribution import list should include the stored import.")
                assert_true(
                    stored_distribution["rows"][0]["traceability"]["source_cells"]["mating_type"] == "B35",
                    "Distribution import should preserve source cell traceability.",
                )
                mice_before_legacy = client.get("/api/mice").json()
                with tempfile.TemporaryDirectory() as legacy_upload_dir:
                    legacy_path = Path(legacy_upload_dir) / "legacy_animal_upload.xlsx"
                    legacy_wb = Workbook()
                    legacy_ws = legacy_wb.active
                    legacy_ws.title = "animal sheet"
                    legacy_ws.append(["Cage No.", "Strain", "Sex", "I.D", "genotype", "DOB", "Mating date", "Pubs"])
                    legacy_ws.append(["C-014", "ApoM Tg/Tg", "M", "MT321", "Tg/Tg", "2026-01-01", "2026-05-01", ""])
                    legacy_ws.append(["", "", "F1", "9p", "pre_weaning", "2026-05-02", "", "2026-05-02 9p"])
                    legacy_wb.save(legacy_path)
                    with legacy_path.open("rb") as legacy_file:
                        legacy_import = client.post(
                            "/api/legacy-workbook-imports",
                            data={"kind": "animal"},
                            files={
                                "file": (
                                    "legacy_animal_upload.xlsx",
                                    legacy_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                            },
                        )
                assert_true(legacy_import.status_code == 200, f"Could not import legacy workbook: {legacy_import.text}")
                legacy_payload = legacy_import.json()
                assert_true(legacy_payload["boundary"] == "parsed or intermediate result", "Legacy workbook import should stay non-canonical.")
                assert_true(legacy_payload["stored_rows"] == 2, "Legacy workbook import row count is wrong.")
                assert_true(legacy_payload["source_record_id"], "Legacy workbook import should create source evidence.")
                legacy_imports = client.get("/api/legacy-workbook-imports").json()
                stored_legacy = next(
                    (
                        item
                        for item in legacy_imports
                        if item["legacy_import_id"] == legacy_payload["legacy_import_id"]
                    ),
                    None,
                )
                assert_true(stored_legacy is not None, "Legacy workbook import list should include the stored import.")
                assert_true(stored_legacy["workbook_kind"] == "legacy_animal_sheet", "Legacy workbook kind was not preserved.")
                assert_true(
                    stored_legacy["rows"][0]["raw_row"]["source_cells"]["display_id"] == "D2",
                    "Legacy workbook import should preserve source cell traceability.",
                )
                assert_true(
                    client.get("/api/mice").json() == mice_before_legacy,
                    "Legacy workbook import should not write canonical mouse state.",
                )
                created = client.post(
                    "/api/assigned-strains",
                    json={
                        "display_name": "ApoM Tg/Tg",
                        "aliases": ["ApoMtg/tg", "ApoM"],
                        "source_type": "manual",
                    },
                )
                assert_true(created.status_code == 200, "Could not create assigned strain scope.")
                payload = created.json()
                assert_true(payload["active"] is True, "Created assigned strain should be active.")
                assert_true("ApoMtg/tg" in payload["aliases"], "Assigned strain aliases were not preserved.")

                rows = client.get("/api/assigned-strains").json()
                assert_true(len(rows) == 1, "Assigned strain list did not return the created scope.")
                assert_true(rows[0]["display_name"] == "ApoM Tg/Tg", "Assigned strain display name changed.")

                deactivated = client.post(f"/api/assigned-strains/{payload['assigned_strain_id']}/deactivate")
                assert_true(deactivated.status_code == 200, "Could not deactivate assigned strain scope.")
                rows = client.get("/api/assigned-strains").json()
                assert_true(rows[0]["active"] is False, "Deactivated assigned strain stayed active.")

                mice_before_distribution = client.get("/api/mice").json()
                distribution_fixture = json.loads(
                    (ROOT / "fixtures" / "sample_distribution_import.json").read_text(encoding="utf-8")
                )
                distribution_import = client.post("/api/distribution-imports", json=distribution_fixture)
                assert_true(distribution_import.status_code == 200, "Could not import distribution assignment JSON.")
                distribution_payload = distribution_import.json()
                assert_true(
                    distribution_payload["boundary"] == "parsed or intermediate result",
                    "Distribution import should stay non-canonical.",
                )
                assert_true(distribution_payload["stored_rows"] >= 3, "Distribution import stored too few rows.")
                imports = client.get("/api/distribution-imports").json()
                imported_fixture = next(
                    (
                        item
                        for item in imports
                        if item["distribution_import_id"] == distribution_payload["distribution_import_id"]
                    ),
                    None,
                )
                assert_true(imported_fixture is not None, f"Distribution import list did not return the new import: {imports!r}")
                assert_true(
                    any(row["mating_type_raw"] == "GFAP Cre; S1PR1 fl/fl" for row in imported_fixture["rows"]),
                    "Distribution assignment rows did not preserve candidate mating type.",
                )
                assert_true(
                    client.get("/api/mice").json() == mice_before_distribution,
                    "Distribution import must not create canonical mouse records.",
                )

                with db.connection() as conn:
                    action_count = conn.execute(
                        "SELECT COUNT(*) AS count FROM action_log WHERE target_id = ?",
                        (payload["assigned_strain_id"],),
                    ).fetchone()["count"]
                assert_true(action_count == 2, "Assigned strain changes should be logged.")

                outside_import = client.post("/api/fixtures/import-sample")
                assert_true(outside_import.status_code == 200, "Could not import fixture without active assigned scope.")
                outside_reviews = client.get("/api/review-items").json()
                assert_true(
                    any(item["issue"] == "Outside assigned strain scope" for item in outside_reviews),
                    "Fixture import without active scope should create outside-scope review items.",
                )

                client.post(
                    "/api/assigned-strains",
                    json={
                        "display_name": "ApoM Tg/Tg",
                        "aliases": ["ApoMtg/tg", "ApoM"],
                        "source_type": "manual",
                    },
                )
                imported = client.post("/api/fixtures/import-sample")
                assert_true(imported.status_code == 200, "Could not import sample fixture through local API.")
                imported_payload = imported.json()
                assert_true(
                    imported_payload["created_or_updated_note_items"] >= 10,
                    "Fixture import should persist parsed note item evidence.",
                )
                assert_true(
                    imported_payload["created_or_updated_mouse_candidates"] >= 3,
                    "Fixture import should create safe mouse candidates from accepted separated rows.",
                )
                mouse_events = client.get("/api/mouse-events").json()
                assert_true(len(mouse_events) >= 3, "Mouse candidates should create mouse event history.")
                assert_true(
                    any(event["event_type"] == "note_added" for event in mouse_events),
                    "Mouse event history should include source-backed note events.",
                )
                review_items = client.get("/api/review-items").json()
                assert_true(len(review_items) >= 4, "Fixture import should create review and validation items.")
                assert_true(
                    any(item["evidence_preview"] and item["note_line_count"] >= 1 for item in review_items),
                    "Review items should expose source note evidence context.",
                )
                assert_true(
                    any(item["source_name"] or item["photo_id"] or item["original_filename"] for item in review_items),
                    "Review items should expose source record or photo context.",
                )
                partial_correction = client.post(
                    f"/api/review-items/{review_items[0]['review_id']}/resolve",
                    json={
                        "resolution_note": "Incomplete correction payload should not be accepted.",
                        "correction_entity_type": "review_item",
                    },
                )
                assert_true(partial_correction.status_code == 400, "Partial review correction payload should be rejected.")
                assert_true(
                    not any(item["issue"] == "Outside assigned strain scope" for item in review_items),
                    "Assigned ApoM scope should clear stale outside-scope review items for now-accepted rows.",
                )
                assert_true(
                    any(item["parse_id"] == "FIXTURE-COUNT-MISMATCH" and item["issue"] == "Count mismatch" for item in review_items),
                    "Count mismatch fixture should create a backend review item.",
                )
                assert_true(
                    any(item["parse_id"] == "FIXTURE-DUPLICATE-ACTIVE" and item["issue"] == "Duplicate active mouse" for item in review_items),
                    "Duplicate active fixture should create a backend review item.",
                )
                assert_true(
                    any(item["parse_id"] == "FIXTURE-EAR-LABEL-CHECK" and item["issue"] == "Ear label needs review" for item in review_items),
                    "Ambiguous ear label fixture should create a note-level review item.",
                )
                count_review = next(item for item in review_items if item["parse_id"] == "FIXTURE-COUNT-MISMATCH")
                resolved = client.post(
                    f"/api/review-items/{count_review['review_id']}/resolve",
                    json={
                        "resolution_note": "Reviewed count mismatch in source note lines.",
                        "resolved_value": count_review["suggested_value"],
                        "correction_entity_type": "review_item",
                        "correction_entity_id": count_review["review_id"],
                        "correction_field_name": "mouse_count",
                        "correction_before_value": count_review["current_value"],
                        "correction_after_value": count_review["suggested_value"],
                    },
                )
                assert_true(resolved.status_code == 200, "Could not resolve review item.")
                resolved_payload = resolved.json()
                assert_true(resolved_payload["correction_id"], "Review resolution with correction values should create a correction log entry.")
                resolved_items = client.get("/api/review-items").json()
                resolved_count_review = next(item for item in resolved_items if item["review_id"] == count_review["review_id"])
                assert_true(resolved_count_review["status"] == "resolved", "Resolved review item stayed open.")
                with db.connection() as conn:
                    review_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'review_resolved' AND target_id = ?
                        """,
                        (count_review["review_id"],),
                    ).fetchone()["count"]
                    correction_row = conn.execute(
                        """
                        SELECT correction_id, entity_type, entity_id, field_name,
                               before_value, after_value, reason, review_id
                        FROM correction_log
                        WHERE correction_id = ?
                        """,
                        (resolved_payload["correction_id"],),
                    ).fetchone()
                    correction_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'correction_applied'
                          AND target_id = ?
                        """,
                        (count_review["review_id"],),
                    ).fetchone()["count"]
                assert_true(review_action_count == 1, "Review resolution should create an action log entry.")
                assert_true(
                    correction_row is not None
                    and correction_row["before_value"] == count_review["current_value"]
                    and correction_row["after_value"] == count_review["suggested_value"]
                    and correction_row["review_id"] == count_review["review_id"],
                    "Review resolution correction should preserve before/after values and review linkage.",
                )
                assert_true(correction_action_count == 1, "Review resolution correction should create a correction action log entry.")
                duplicate_resolve = client.post(
                    f"/api/review-items/{count_review['review_id']}/resolve",
                    json={"resolution_note": "Duplicate review resolution should be blocked."},
                )
                assert_true(duplicate_resolve.status_code == 409, "Resolved review items should not be resolved again.")
                duplicate_active_review = next(item for item in resolved_items if item["parse_id"] == "FIXTURE-DUPLICATE-ACTIVE")
                partial_correction = client.post(
                    f"/api/review-items/{duplicate_active_review['review_id']}/resolve",
                    json={
                        "resolution_note": "Incomplete correction metadata should be rejected.",
                        "correction_entity_type": "review_item",
                    },
                )
                assert_true(partial_correction.status_code == 400, "Partial correction metadata should not resolve a review item.")
                still_open_reviews = client.get("/api/review-items").json()
                assert_true(
                    any(
                        item["review_id"] == duplicate_active_review["review_id"]
                        and item["status"] == "open"
                        for item in still_open_reviews
                    ),
                    "Rejected partial correction should leave the review item open.",
                )
                note_items = client.get("/api/note-items").json()
                mice = client.get("/api/mice").json()
                assert_true(
                    any(item["raw_line_text"] == "MT321 R'" and item["parsed_ear_label_code"] == "R_PRIME" for item in note_items),
                    "Note item API should expose parsed ear label evidence.",
                )
                assert_true(
                    any(
                        item["raw_line_text"] == "MT399 R0"
                        and item["parsed_ear_label_code"] == "R_CIRCLE"
                        and item["parsed_ear_label_review_status"] == "check"
                        and item["needs_review"] == 1
                        for item in note_items
                    ),
                    "Ambiguous ear label note should stay reviewable with raw evidence.",
                )
                assert_true(
                    any(
                        mouse["display_id"] == "MT399"
                        and mouse["ear_label_raw"] == "R0"
                        and mouse["ear_label_code"] is None
                        and mouse["ear_label_review_status"] == "check"
                        for mouse in mice
                    ),
                    "Mouse candidate should not accept an uncertain normalized ear label.",
                )
                assert_true(
                    any(mouse["display_id"] == "MT323" and mouse["status"] == "moved" for mouse in mice),
                    "Mouse API should expose moved candidate from single-struck note line.",
                )
                cage = client.post(
                    "/api/cages",
                    json={"cage_label": "C-014", "location": "Room A / Rack 2", "cage_type": "holding"},
                )
                assert_true(cage.status_code == 200, "Could not create cage registry entry.")
                cage_payload = cage.json()
                assert_true(cage_payload["source_record_id"], "Cage creation should create source evidence.")
                moved_mouse = next(mouse for mouse in mice if mouse["display_id"] == "MT321")
                moved_to_cage = client.post(
                    f"/api/mice/{moved_mouse['mouse_id']}/move-cage",
                    json={"cage_id": cage_payload["cage_id"], "note": "Verified cage movement flow."},
                )
                assert_true(moved_to_cage.status_code == 200, "Could not move mouse to cage.")
                moved_payload = moved_to_cage.json()
                assert_true(moved_payload["event_id"], "Cage move should create a mouse event.")
                cage_rows = client.get("/api/cages").json()
                assert_true(
                    any(row["cage_label"] == "C-014" and row["active_mouse_count"] == 1 for row in cage_rows),
                    "Cage list should show active mouse count after assignment.",
                )
                mice_after_cage = client.get("/api/mice", params={"query": "C-014"}).json()
                assert_true(
                    any(mouse["display_id"] == "MT321" and mouse["current_cage_label"] == "C-014" for mouse in mice_after_cage),
                    "Mouse search should find current cage assignment.",
                )
                cage_events = client.get("/api/mouse-events").json()
                assert_true(
                    any(event["event_type"] == "moved" and event["related_entity_id"] == cage_payload["cage_id"] for event in cage_events),
                    "Cage movement should be present in mouse event history.",
                )
                audit_trace = client.get(f"/api/mice/{moved_mouse['mouse_id']}/audit-trace")
                assert_true(audit_trace.status_code == 200, "Could not load mouse audit trace.")
                audit_payload = audit_trace.json()
                assert_true(audit_payload["source_layer"] == "export or view", "Audit trace should stay in the export/view layer.")
                assert_true(audit_payload["mouse"]["display_id"] == "MT321", "Audit trace should return the requested mouse.")
                audit_categories = {item["category"] for item in audit_payload["timeline"]}
                assert_true({"note_line", "mouse_event"}.issubset(audit_categories), "Audit trace should include source note lines and mouse events.")
                assert_true(
                    any(event["event_type"] == "moved" for event in audit_payload["events"]),
                    "Audit trace should include cage movement events.",
                )
                filtered_mice = client.get("/api/mice", params={"query": "MT321"}).json()
                assert_true(
                    filtered_mice and all("MT321" in mouse["display_id"] for mouse in filtered_mice),
                    "Mouse search filter should narrow candidate records by display ID.",
                )
                search_payload = client.get("/api/search", params={"query": "PV-Cre"}).json()
                assert_true(search_payload["query"] == "PV-Cre", "Search endpoint should echo the active query.")
                assert_true(
                    any(strain["strain_name"] == "PV-Cre" for strain in search_payload["strains"]),
                    "Search endpoint should include strain registry matches.",
                )
                csv_response = client.get("/api/exports/mice.csv", params={"query": "MT321"})
                assert_true(csv_response.status_code == 200, "Mouse CSV export endpoint failed.")
                assert_true(
                    csv_response.headers["content-type"].startswith("text/csv"),
                    "Mouse CSV export should return CSV content.",
                )
                csv_text = csv_response.text
                assert_true("display_id" in csv_text and "MT321" in csv_text, "Mouse CSV export is missing expected rows.")
                assert_true("MT323" not in csv_text, "Filtered mouse CSV export should exclude non-matching mice.")
                export_log = client.get("/api/export-log").json()
                assert_true(export_log, "CSV generation should create an export log entry.")
                assert_true(export_log[0]["export_type"] == "mouse_csv", "Export log should record CSV export type.")
                assert_true(export_log[0]["filename"] == "mouse_records_filtered.csv", "Export log should preserve generated filename.")
                assert_true(export_log[0]["query"] == "MT321", "Export log should preserve the export filter.")
                assert_true(export_log[0]["row_count"] == len(filtered_mice), "Export log should preserve exported row count.")
                assert_true(export_log[0]["source_layer"] == "export or view", "Export log should stay in the export/view layer.")
                blocked_export = client.get("/api/exports/mice.csv", params={"query": "MT321", "require_ready": "true"})
                assert_true(blocked_export.status_code == 409, "Final CSV export should be blocked by open review items.")
                blocked_separation_xlsx = client.get("/api/exports/separation.xlsx")
                assert_true(blocked_separation_xlsx.status_code == 409, "Final separation XLSX export should be blocked by open review items.")
                blocked_animal_xlsx = client.get("/api/exports/animal-sheet.xlsx")
                assert_true(blocked_animal_xlsx.status_code == 409, "Final animal sheet XLSX export should be blocked by open review items.")
                blocked_payload = blocked_export.json()["detail"]
                assert_true(
                    blocked_payload["source_layer"] == "export or view" and blocked_payload["blocked_review_count"] > 0,
                    "Blocked final export should report export/view layer and blocker count.",
                )
                assert_true(
                    blocked_payload["review_blockers"]
                    and {
                        "review_id",
                        "issue",
                        "severity",
                        "review_reason",
                        "evidence_preview",
                        "note_line_count",
                    }.issubset(blocked_payload["review_blockers"][0]),
                    "Blocked final CSV export should include actionable review blocker details.",
                )
                blocked_separation_payload = blocked_separation_xlsx.json()["detail"]
                blocked_animal_payload = blocked_animal_xlsx.json()["detail"]
                assert_true(
                    blocked_separation_payload["review_blockers"]
                    and blocked_animal_payload["review_blockers"],
                    "Blocked workbook exports should include review blocker previews.",
                )
                blocked_logs = client.get("/api/export-log").json()
                blocked_log = next(
                    (item for item in blocked_logs if item["export_type"] == "mouse_csv" and item["status"] == "blocked"),
                    None,
                )
                assert_true(blocked_log is not None, "Blocked final CSV export should create a blocked export log entry.")
                assert_true(blocked_log["status"] == "blocked", "Blocked final export should create a blocked export log entry.")
                assert_true(blocked_log["filename"] == "mouse_records_filtered.csv", "Blocked export log should preserve intended filename.")
                export_preview = client.get("/api/export-preview").json()
                assert_true(export_preview["source_layer"] == "export or view", "Export preview should stay an export/view layer.")
                assert_true(export_preview["export_type"] == "separation_preview", "Export preview should identify its workbook-like shape.")
                assert_true(export_preview["preview_row_count"] >= 3, "Export preview should include mouse candidate rows.")
                assert_true(
                    export_preview["latest_data_change_at"] and export_preview["latest_generated_export_at"],
                    "Export preview should expose data and generated export timestamps.",
                )
                assert_true(export_preview["export_stale"] is False, "Freshly generated CSV export should not be stale before more data changes.")
                assert_true(
                    export_preview["expected_separation_filename"].endswith("분리 현황표.xlsx"),
                    "Export preview should expose the expected separation workbook filename.",
                )
                assert_true(
                    export_preview["expected_animal_sheet_filename"].endswith("animal sheet.xlsx"),
                    "Export preview should expose the expected animal sheet workbook filename.",
                )
                assert_true(
                    export_preview["separation_columns"][:5] == ["Cage number", "Strain", "Genotype", "total", "DOB"],
                    "Separation preview should expose senior-workbook-style column labels.",
                )
                assert_true(
                    any(row["total"].endswith("p") and row["source_note_item_ids"] for row in export_preview["separation_rows"]),
                    "Separation preview should group mouse records into workbook-like sex/count rows with traceability.",
                )
                assert_true(
                    any(row["display_id"] == "MT321" and row["source_note_item_id"] for row in export_preview["preview_rows"]),
                    "Export preview rows should preserve source note traceability.",
                )
                assert_true(
                    export_preview["blocked_review_items"] >= len(export_preview["review_blockers"]),
                    "Export preview should expose review blocker details up to its display limit.",
                )
                assert_true(export_preview["ready"] is False, "Open review blockers should keep export preview blocked.")
                dashboard_before = client.get("/api/genotyping-dashboard")
                assert_true(dashboard_before.status_code == 200, "Genotyping dashboard endpoint failed.")
                dashboard_before_rows = {card["key"]: card["count"] for card in dashboard_before.json()}
                assert_true(
                    dashboard_before_rows.get("not_sampled", 0) >= 1,
                    "Genotyping dashboard should count newly separated mice that need sampling.",
                )
                genotyping_target = next(mouse for mouse in mice if mouse["display_id"] == "MT321")
                audit_trace = client.get(f"/api/mice/{genotyping_target['mouse_id']}/audit-trace")
                assert_true(audit_trace.status_code == 200, "Mouse audit trace endpoint failed.")
                audit_payload = audit_trace.json()
                assert_true(
                    audit_payload["source_layer"] == "export or view"
                    and audit_payload["mouse"]["mouse_id"] == genotyping_target["mouse_id"],
                    "Mouse audit trace should identify the selected mouse and boundary.",
                )
                assert_true(
                    any(item["category"] in {"note_line", "mouse_event", "review"} for item in audit_payload["timeline"]),
                    "Mouse audit trace should combine note, event, or review evidence in one timeline.",
                )
                target_rule = client.post(
                    "/api/strain-target-genotypes",
                    json={
                        "strain_text": genotyping_target["raw_strain_text"],
                        "target_genotype": "Tg/Tg",
                        "purpose": "mating_candidate",
                    },
                )
                assert_true(target_rule.status_code == 200, "Could not create configurable target genotype rule.")
                target_rules = client.get("/api/strain-target-genotypes").json()
                assert_true(
                    any(
                        rule["strain_text"] == genotyping_target["raw_strain_text"]
                        and rule["target_genotype"] == "Tg/Tg"
                        and rule["purpose"] == "mating_candidate"
                        for rule in target_rules
                    ),
                    "Target genotype rule API should preserve configured rule values.",
                )
                cage = client.post(
                    "/api/cages",
                    json={
                        "cage_label": "A-101",
                        "location": "Room A",
                        "rack": "R1",
                        "shelf": "S2",
                        "cage_type": "holding",
                        "note": "Verification cage.",
                    },
                )
                assert_true(cage.status_code == 200, "Could not create cage registry entry.")
                cage_payload = cage.json()
                assert_true(cage_payload["source_record_id"], "Cage creation should leave source evidence.")
                duplicate_cage = client.post("/api/cages", json={"cage_label": "a-101"})
                assert_true(duplicate_cage.status_code == 409, "Duplicate cage labels should be rejected case-insensitively.")
                cage_move = client.post(
                    f"/api/mice/{genotyping_target['mouse_id']}/move-cage",
                    json={"cage_id": cage_payload["cage_id"], "note": "Moved for verification."},
                )
                assert_true(cage_move.status_code == 200, "Could not assign mouse to cage.")
                cages = client.get("/api/cages").json()
                assert_true(
                    any(item["cage_label"] == "A-101" and item["active_mouse_count"] == 1 for item in cages),
                    "Cage list should include active assignment counts.",
                )
                female_parent = next(mouse for mouse in mice if mouse["display_id"] == "MT322")
                mating = client.post(
                    "/api/matings",
                    json={
                        "mating_label": "MT321 x MT322",
                        "male_mouse_id": genotyping_target["mouse_id"],
                        "female_mouse_id": female_parent["mouse_id"],
                        "strain_goal": genotyping_target["raw_strain_text"],
                        "expected_genotype": "Tg/Tg",
                        "start_date": "2026-05-01",
                        "purpose": "verification",
                    },
                )
                assert_true(mating.status_code == 200, "Could not create mating registry entry.")
                mating_payload = mating.json()
                mating_rows = client.get("/api/matings").json()
                assert_true(
                    any(
                        row["mating_id"] == mating_payload["mating_id"]
                        and "MT321" in row["male_mice"]
                        and "MT322" in row["female_mice"]
                        for row in mating_rows
                    ),
                    "Mating list should expose linked parent mice by role.",
                )
                missing_parent_mating = client.post(
                    "/api/matings",
                    json={"mating_label": "bad mating", "male_mouse_id": "missing_mouse"},
                )
                assert_true(missing_parent_mating.status_code == 404, "Mating creation should reject missing parent mouse IDs.")
                litter = client.post(
                    "/api/litters",
                    json={
                        "litter_label": "L-MT321-001",
                        "mating_id": mating_payload["mating_id"],
                        "birth_date": "2026-05-02",
                        "number_born": 10,
                        "number_alive": 9,
                        "status": "born",
                    },
                )
                assert_true(litter.status_code == 200, "Could not create litter registry entry.")
                litter_payload = litter.json()
                litter_rows = client.get("/api/litters").json()
                assert_true(
                    any(row["litter_id"] == litter_payload["litter_id"] and row["number_born"] == 10 for row in litter_rows),
                    "Litter list should expose source-backed litter counts.",
                )
                offspring = client.post(
                    f"/api/litters/{litter_payload['litter_id']}/offspring",
                    json={
                        "count": 2,
                        "display_prefix": "MT321-L1",
                        "start_number": 1,
                        "sex": "unknown",
                        "cage_id": cage_payload["cage_id"],
                        "note": "Generated from reviewed litter count.",
                    },
                )
                assert_true(offspring.status_code == 200, "Could not create offspring mouse records from litter.")
                offspring_payload = offspring.json()
                assert_true(offspring_payload["created_count"] == 2, "Offspring creation should return created count.")
                assert_true(offspring_payload["source_record_id"], "Offspring creation should preserve source evidence.")
                duplicate_offspring = client.post(
                    f"/api/litters/{litter_payload['litter_id']}/offspring",
                    json={"count": 1, "display_prefix": "MT321-L1", "start_number": 1},
                )
                assert_true(duplicate_offspring.status_code == 409, "Duplicate offspring IDs should be rejected.")
                offspring_rows = client.get("/api/mice", params={"query": "MT321-L1"}).json()
                assert_true(len(offspring_rows) == 2, "Mouse search should include generated offspring records.")
                assert_true(
                    all(
                        row["father_id"] == genotyping_target["mouse_id"]
                        and row["mother_id"] == female_parent["mouse_id"]
                        and row["litter_id"] == litter_payload["litter_id"]
                        and row["source_record_id"] == offspring_payload["source_record_id"]
                        and row["current_cage_label"] == "A-101"
                        for row in offspring_rows
                    ),
                    "Generated offspring should preserve parent, litter, source, and cage traceability.",
                )
                litter_rows_after_offspring = client.get("/api/litters").json()
                assert_true(
                    any(row["litter_id"] == litter_payload["litter_id"] and row["offspring_count"] == 2 for row in litter_rows_after_offspring),
                    "Litter list should expose generated offspring count.",
                )
                animal_preview = client.get("/api/export-preview").json()
                assert_true(
                    animal_preview["animal_sheet_columns"][:8]
                    == ["Cage No.", "Strain", "Sex", "I.D", "genotype", "DOB", "Mating date", "Pubs"],
                    "Animal sheet preview should expose mating-workbook-style column labels.",
                )
                assert_true(
                    any(
                        row["cage_no"] == "1"
                        and row["strain"] == genotyping_target["raw_strain_text"]
                        and row["mating_date"] == "2026-05-01"
                        for row in animal_preview["animal_sheet_rows"]
                    ),
                    "Animal sheet preview should include parent rows grouped by mating cage.",
                )
                assert_true(
                    any(row["sex"] == "F1" and row["mouse_id"] == "9p" and row["status"] == "pre_weaning" for row in animal_preview["animal_sheet_rows"]),
                    "Animal sheet preview should include litter rows with pup counts and status.",
                )
                over_weaned = client.post(
                    f"/api/litters/{litter_payload['litter_id']}/wean",
                    json={"weaning_date": "2026-05-23", "number_weaned": 3},
                )
                assert_true(over_weaned.status_code == 409, "Weaning should reject counts above generated offspring records.")
                weaned = client.post(
                    f"/api/litters/{litter_payload['litter_id']}/wean",
                    json={
                        "weaning_date": "2026-05-23",
                        "number_weaned": 2,
                        "note": "Verified weaning from reviewed litter card.",
                    },
                )
                assert_true(weaned.status_code == 200, "Could not complete litter weaning.")
                weaned_payload = weaned.json()
                assert_true(weaned_payload["status"] == "weaned", "Weaning should mark litter status as weaned.")
                assert_true(weaned_payload["number_weaned"] == 2, "Weaning should preserve reviewed weaned count.")
                assert_true(weaned_payload["source_record_id"], "Weaning should preserve source evidence.")
                duplicate_wean = client.post(
                    f"/api/litters/{litter_payload['litter_id']}/wean",
                    json={"weaning_date": "2026-05-24", "number_weaned": 2},
                )
                assert_true(duplicate_wean.status_code == 409, "Already-weaned litters should not be silently overwritten.")
                weaned_offspring_rows = client.get("/api/mice", params={"query": "MT321-L1"}).json()
                assert_true(
                    all(row["status"] == "active" and row["next_action"] == "sample_needed" for row in weaned_offspring_rows),
                    "Weaned offspring should move from weaning pending to active sample-needed workflow.",
                )
                requested_genotyping = client.post(
                    "/api/genotyping/request",
                    json={
                        "mouse_id": weaned_offspring_rows[0]["mouse_id"],
                        "sample_id": "TAIL-MT321-L1-01",
                        "target_name": "ApoM Tg/Tg",
                        "note": "Requested after weaning verification.",
                    },
                )
                assert_true(requested_genotyping.status_code == 200, "Could not request genotyping for weaned offspring.")
                requested_payload = requested_genotyping.json()
                assert_true(requested_payload["sample_id"] == "TAIL-MT321-L1-01", "Genotyping request should preserve sample ID.")
                assert_true(requested_payload["genotyping_status"] == "submitted", "Genotyping request should mark mouse submitted.")
                assert_true(requested_payload["next_action"] == "awaiting_result", "Genotyping request should move mouse to awaiting result.")
                litter_rows_after_weaning = client.get("/api/litters").json()
                assert_true(
                    any(
                        row["litter_id"] == litter_payload["litter_id"]
                        and row["status"] == "weaned"
                        and row["number_weaned"] == 2
                        and row["weaning_date"] == "2026-05-23"
                        for row in litter_rows_after_weaning
                    ),
                    "Litter list should expose completed weaning state.",
                )
                genotyping_update = client.post(
                    "/api/genotyping/update",
                    json={
                        "mouse_id": genotyping_target["mouse_id"],
                        "sample_id": "MT321",
                        "raw_result": "Tg/Tg",
                        "normalized_result": "Tg/Tg",
                    },
                )
                assert_true(genotyping_update.status_code == 200, "Could not update genotyping workflow state.")
                genotyping_payload = genotyping_update.json()
                assert_true(genotyping_payload["genotyping_status"] == "resulted", "Genotyping result should mark mouse resulted.")
                assert_true(genotyping_payload["target_match_status"] == "matches_target", "Genotyping result should use configured target matching.")
                assert_true(genotyping_payload["next_action"] == "consider_for_mating", "Matching target genotype should suggest a mating review action.")
                stale_preview = client.get("/api/export-preview").json()
                assert_true(
                    stale_preview["export_stale"] is True
                    and stale_preview["latest_data_change_at"] >= stale_preview["latest_generated_export_at"],
                    "Export preview should warn when mouse data changed after the last generated export.",
                )
                duplicate_resulted_request = client.post(
                    "/api/genotyping/request",
                    json={"mouse_id": genotyping_target["mouse_id"], "sample_id": "already-resulted"},
                )
                assert_true(duplicate_resulted_request.status_code == 409, "Resulted mice should not accept a new genotyping request silently.")
                parent_trace = client.get(f"/api/mice/{genotyping_target['mouse_id']}/audit-trace")
                assert_true(parent_trace.status_code == 200, "Mouse audit trace endpoint should return parent mouse evidence.")
                parent_trace_payload = parent_trace.json()
                parent_categories = {item["category"] for item in parent_trace_payload["timeline"]}
                assert_true(parent_trace_payload["source_layer"] == "export or view", "Mouse audit trace should stay a review/export view.")
                assert_true(parent_trace_payload["note_items"], "Mouse audit trace should include parsed note-line evidence.")
                assert_true("note_line" in parent_categories, "Mouse audit trace timeline should include note-line evidence.")
                assert_true("genotyping" in parent_categories, "Mouse audit trace timeline should include genotyping records.")
                offspring_trace = client.get(f"/api/mice/{weaned_offspring_rows[0]['mouse_id']}/audit-trace")
                assert_true(offspring_trace.status_code == 200, "Mouse audit trace endpoint should return offspring evidence.")
                offspring_trace_payload = offspring_trace.json()
                offspring_categories = {item["category"] for item in offspring_trace_payload["timeline"]}
                assert_true("mouse_event" in offspring_categories, "Offspring audit trace should include born/weaned/request events.")
                assert_true("genotyping" in offspring_categories, "Offspring audit trace should include pending genotyping request.")
                assert_true(
                    offspring_trace_payload["source_records"],
                    "Offspring audit trace should preserve source records linked through mouse events.",
                )
                dashboard_after = {card["key"]: card["count"] for card in client.get("/api/genotyping-dashboard").json()}
                assert_true(
                    dashboard_after.get("target_confirmed", 0) >= 1,
                    "Genotyping dashboard should count mice with confirmed target genotypes.",
                )
                genotyping_records = client.get("/api/genotyping-records").json()
                assert_true(
                    any(record["mouse_id"] == genotyping_target["mouse_id"] and record["normalized_result"] == "Tg/Tg" for record in genotyping_records),
                    "Genotyping record history should preserve the entered result.",
                )
                assert_true(
                    any(record["sample_id"] == "TAIL-MT321-L1-01" and record["result_status"] == "pending" for record in genotyping_records),
                    "Genotyping request should create a pending genotyping record.",
                )
                missing_audit = client.get("/api/mice/mouse-does-not-exist/audit-trail")
                assert_true(missing_audit.status_code == 404, "Missing mouse audit trail should return 404.")
                audit_trace = client.get(f"/api/mice/{genotyping_target['mouse_id']}/audit-trail")
                assert_true(audit_trace.status_code == 200, "Mouse audit trail endpoint failed.")
                audit_payload = audit_trace.json()
                assert_true(audit_payload["source_layer"] == "export or view", "Audit trail should be a read-only export/view layer.")
                assert_true(audit_payload["mouse"]["display_id"] == "MT321", "Audit trail should include the selected mouse.")
                audit_categories = {item["category"] for item in audit_payload["timeline"]}
                assert_true(
                    {"note_line", "mouse_event", "genotyping", "cage_assignment", "action_log"}.issubset(audit_categories),
                    "Audit trail should combine source notes, events, genotyping, cage history, and action logs.",
                )
                assert_true(
                    any(action["action_type"] == "genotyping_resulted" for action in audit_payload["actions"]),
                    "Audit trail should expose direct action log records for the mouse.",
                )
                assert_true(
                    any(assignment["cage_label"] == "C-014" for assignment in audit_payload["cage_assignments"]),
                    "Audit trail should expose cage assignment history.",
                )
                offspring_audit = client.get(f"/api/mice/{requested_payload['mouse_id']}/audit-trail").json()
                assert_true(
                    offspring_audit["lineage"]["litter"]["litter_id"] == litter_payload["litter_id"]
                    and offspring_audit["lineage"]["father"]["mouse_id"] == genotyping_target["mouse_id"],
                    "Offspring audit trail should expose litter and parent lineage.",
                )
                genotyping_export = client.get("/api/exports/genotyping-worklist.csv", params={"query": "MT321"})
                assert_true(genotyping_export.status_code == 200, "Genotyping worklist CSV export endpoint failed.")
                assert_true(
                    "genotyping_worklist_filtered.csv" in genotyping_export.headers.get("content-disposition", ""),
                    "Filtered genotyping worklist should use the filtered filename.",
                )
                assert_true(
                    "target_match_status" in genotyping_export.text
                    and "matches_target" in genotyping_export.text
                    and "consider_for_mating" in genotyping_export.text,
                    "Genotyping worklist export should include target match and next action fields.",
                )
                genotyping_export_log = client.get("/api/export-log").json()
                assert_true(
                    genotyping_export_log[0]["export_type"] == "genotyping_worklist_csv",
                    "Export log should record companion genotyping worklist exports.",
                )
                filtered_mice = client.get("/api/mice", params={"query": "MT321"}).json()
                assert_true(
                    filtered_mice and all("MT321" in mouse["display_id"] for mouse in filtered_mice),
                    "Mouse API query should filter mouse records.",
                )
                search = client.get("/api/search", params={"query": "Tg/Tg"}).json()
                assert_true(
                    any(mouse["display_id"] == "MT321" for mouse in search["mice"]),
                    "Search API should include matching genotyping mouse records.",
                )
                csv_export = client.get("/api/exports/mice.csv", params={"query": "MT321"})
                assert_true(csv_export.status_code == 200, "Mouse CSV export endpoint failed.")
                assert_true(
                    "mouse_records_filtered.csv" in csv_export.headers.get("content-disposition", ""),
                    "Filtered CSV export should use the filtered filename.",
                )
                assert_true(
                    "display_id" in csv_export.text and "MT321" in csv_export.text,
                    "Mouse CSV export should include headers and filtered mouse rows.",
                )
                offspring_csv = client.get("/api/exports/mice.csv", params={"query": "MT321-L1"})
                assert_true(offspring_csv.status_code == 200, "Offspring CSV export endpoint failed.")
                assert_true(
                    "father_id" in offspring_csv.text
                    and "mother_id" in offspring_csv.text
                    and "litter_id" in offspring_csv.text
                    and offspring_payload["source_record_id"] in offspring_csv.text,
                    "Offspring CSV export should include lineage and source traceability fields.",
                )
                with db.connection() as conn:
                    note_count = conn.execute(
                        "SELECT COUNT(*) AS count FROM card_note_item_log"
                    ).fetchone()["count"]
                    mouse_count = conn.execute("SELECT COUNT(*) AS count FROM mouse_master").fetchone()["count"]
                    moved_count = conn.execute(
                        "SELECT COUNT(*) AS count FROM mouse_master WHERE status = 'moved'"
                    ).fetchone()["count"]
                    duplicate_leak_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_master
                        WHERE source_note_item_id LIKE ?
                        """,
                        ("note_FIXTURE-DUPLICATE-ACTIVE_%",),
                    ).fetchone()["count"]
                    genotyping_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'genotyping_resulted' AND target_id = ?
                        """,
                        (genotyping_target["mouse_id"],),
                    ).fetchone()["count"]
                    cage_move_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'mouse_cage_moved' AND target_id = ?
                        """,
                        (genotyping_target["mouse_id"],),
                    ).fetchone()["count"]
                    cage_move_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'moved' AND related_entity_type = 'cage'
                        """
                    ).fetchone()["count"]
                    active_cage_assignment_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_cage_assignment
                        WHERE mouse_id = ? AND status = 'active'
                        """,
                        (genotyping_target["mouse_id"],),
                    ).fetchone()["count"]
                    ended_cage_assignment_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_cage_assignment
                        WHERE mouse_id = ? AND status = 'ended'
                        """,
                        (genotyping_target["mouse_id"],),
                    ).fetchone()["count"]
                    mating_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'mating_created' AND target_id = ?
                        """,
                        (mating_payload["mating_id"],),
                    ).fetchone()["count"]
                    litter_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'litter_created' AND target_id = ?
                        """,
                        (litter_payload["litter_id"],),
                    ).fetchone()["count"]
                    paired_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'paired' AND related_entity_type = 'mating'
                        """
                    ).fetchone()["count"]
                    litter_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'litter_produced' AND related_entity_type = 'litter'
                        """
                    ).fetchone()["count"]
                    offspring_born_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'born' AND related_entity_type = 'litter'
                        """
                    ).fetchone()["count"]
                    offspring_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'offspring_created' AND target_id = ?
                        """,
                        (litter_payload["litter_id"],),
                    ).fetchone()["count"]
                    weaned_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'weaned' AND related_entity_type = 'litter'
                        """
                    ).fetchone()["count"]
                    weaned_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'litter_weaned' AND target_id = ?
                        """,
                        (litter_payload["litter_id"],),
                    ).fetchone()["count"]
                    genotyping_request_action_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM action_log
                        WHERE action_type = 'genotyping_requested'
                        """
                    ).fetchone()["count"]
                    tail_biopsy_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'tail_biopsy' AND related_entity_type = 'genotyping_record'
                        """
                    ).fetchone()["count"]
                    genotyping_request_event_count = conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM mouse_event
                        WHERE event_type = 'genotyping_requested' AND related_entity_type = 'genotyping_record'
                        """
                    ).fetchone()["count"]
                assert_true(note_count >= 10, "Persisted note item evidence count is too low.")
                assert_true(mouse_count >= 5, "Persisted mouse candidate count is too low.")
                assert_true(moved_count >= 1, "Single-struck mouse note should create a moved candidate.")
                assert_true(duplicate_leak_count == 0, "Duplicate active fixture should not create mouse candidates.")
                assert_true(genotyping_action_count == 1, "Genotyping update should create an action log entry.")
                assert_true(cage_move_action_count >= 2, "Each cage move should create an action log entry.")
                assert_true(cage_move_event_count >= 1, "Cage move should create a mouse event.")
                assert_true(active_cage_assignment_count == 1, "Cage moves should leave only one active assignment per mouse.")
                assert_true(ended_cage_assignment_count >= 1, "Cage moves should close the previous active assignment.")
                assert_true(mating_action_count == 1, "Mating creation should create an action log entry.")
                assert_true(litter_action_count == 1, "Litter creation should create an action log entry.")
                assert_true(paired_event_count >= 2, "Mating creation should create parent pairing events.")
                assert_true(litter_event_count >= 2, "Litter creation should create parent litter events.")
                assert_true(offspring_born_event_count == 2, "Offspring creation should create one birth event per mouse.")
                assert_true(offspring_action_count == 1, "Offspring creation should create an action log entry.")
                assert_true(weaned_event_count == 2, "Weaning should create one event per weaned offspring.")
                assert_true(weaned_action_count == 1, "Weaning should create an action log entry.")
                assert_true(genotyping_request_action_count == 1, "Genotyping request should create an action log entry.")
                assert_true(tail_biopsy_event_count == 1, "Genotyping request should create a tail biopsy event.")
                assert_true(genotyping_request_event_count == 1, "Genotyping request should create a request event.")
                correction = client.post(
                    "/api/corrections",
                    json={
                        "entity_type": "strain",
                        "entity_id": strain_payload["strain_id"],
                        "field_name": "common_name",
                        "before_value": "",
                        "after_value": "PV-Cre line",
                        "reason": "Verified local correction workflow.",
                        "source_record_id": strain_payload["source_record_id"],
                    },
                )
                assert_true(correction.status_code == 200, "Could not create correction log entry.")
                corrections = client.get("/api/corrections").json()
                assert_true(
                    corrections[0]["before_value"] == "" and corrections[0]["after_value"] == "PV-Cre line",
                    "Correction log should preserve before and after values.",
                )
                remaining_reviews = [
                    item for item in client.get("/api/review-items").json() if item["status"] == "open"
                ]
                assert_true(remaining_reviews, "Fixture should still have open review blockers before final release.")
                for item in remaining_reviews:
                    release_review = client.post(
                        f"/api/review-items/{item['review_id']}/resolve",
                        json={
                            "resolution_note": "Verified blocker before releasing ready CSV export.",
                            "resolved_value": item.get("suggested_value") or item.get("current_value") or "",
                        },
                    )
                    assert_true(
                        release_review.status_code == 200,
                        f"Could not resolve remaining review blocker: {release_review.status_code} {release_review.text}",
                    )
                ready_preview = client.get("/api/export-preview").json()
                assert_true(ready_preview["ready"] is True, "Resolved review blockers should make export preview ready.")
                assert_true(ready_preview["blocked_review_items"] == 0, "Ready export preview should have no blockers.")
                ready_export = client.get("/api/exports/mice.csv", params={"query": "MT321", "require_ready": "true"})
                assert_true(ready_export.status_code == 200, "Ready CSV export should succeed after review resolution.")
                assert_true("MT321" in ready_export.text, "Ready CSV export should include the filtered mouse row.")
                post_ready_preview = client.get("/api/export-preview").json()
                assert_true(post_ready_preview["export_stale"] is False, "Generated ready export should clear the stale export warning.")
                ready_separation_xlsx = client.get("/api/exports/separation.xlsx")
                assert_true(ready_separation_xlsx.status_code == 200, "Ready separation XLSX export should succeed after review resolution.")
                assert_true(
                    ready_separation_xlsx.content[:4] == b"PK\x03\x04",
                    "Separation XLSX export should be a ZIP-based workbook.",
                )
                separation_disposition = ready_separation_xlsx.headers.get("content-disposition", "")
                assert_true(
                    "filename*=UTF-8''" in separation_disposition and "separation.xlsx" in separation_disposition,
                    "Separation XLSX export should expose a safe fallback and UTF-8 filename.",
                )
                with zipfile.ZipFile(io.BytesIO(ready_separation_xlsx.content)) as workbook_zip:
                    assert_true(
                        "xl/workbook.xml" in workbook_zip.namelist()
                        and "xl/styles.xml" in workbook_zip.namelist()
                        and "xl/worksheets/sheet1.xml" in workbook_zip.namelist()
                        and "xl/worksheets/sheet2.xml" in workbook_zip.namelist(),
                        "Separation XLSX should contain the required workbook parts.",
                    )
                    separation_workbook_xml = workbook_zip.read("xl/workbook.xml").decode("utf-8")
                    separation_sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")
                    separation_trace_xml = workbook_zip.read("xl/worksheets/sheet2.xml").decode("utf-8")
                    separation_styles_xml = workbook_zip.read("xl/styles.xml").decode("utf-8")
                assert_true("분리 현황표" in separation_workbook_xml, "Separation XLSX should use the lab workbook sheet name.")
                assert_true("Sampling point" in separation_sheet_xml, "Separation XLSX should include the template header.")
                assert_true("ApoM Tg/Tg" in separation_sheet_xml, "Separation XLSX should include accepted strain rows.")
                assert_true("<cols>" in separation_sheet_xml and 's="1"' in separation_sheet_xml, "Separation XLSX should include column widths and styled headers.")
                assert_true("Export_Trace" in separation_workbook_xml and "Source note" in separation_trace_xml, "Separation XLSX should include traceability sheet.")
                assert_true("<b/>" in separation_styles_xml, "Separation XLSX should include bold header style.")
                assert_true(load_workbook is not None, "openpyxl is required to validate generated XLSX workbooks.")
                separation_workbook = load_workbook(io.BytesIO(ready_separation_xlsx.content), data_only=True)
                assert_true("분리 현황표" in separation_workbook.sheetnames, "openpyxl should load the separation sheet name.")
                assert_true("Export_Trace" in separation_workbook.sheetnames, "openpyxl should load the separation trace sheet.")
                separation_sheet = separation_workbook["분리 현황표"]
                assert_true(
                    [separation_sheet.cell(1, column).value for column in range(1, 9)]
                    == ["Cage number", "Strain", "Genotype", "total", "DOB", "WT", "Tg", "Sampling point"],
                    "openpyxl should read the separation workbook headers.",
                )
                assert_true(
                    any(row[1] == "ApoM Tg/Tg" for row in separation_sheet.iter_rows(min_row=2, values_only=True)),
                    "openpyxl should read accepted separation strain rows.",
                )
                separation_trace_sheet = separation_workbook["Export_Trace"]
                assert_true(
                    separation_trace_sheet.cell(1, 2).value == "Source note"
                    and separation_trace_sheet.cell(1, 4).value == "Boundary",
                    "openpyxl should read separation trace headers.",
                )
                ready_animal_xlsx = client.get("/api/exports/animal-sheet.xlsx")
                assert_true(ready_animal_xlsx.status_code == 200, "Ready animal sheet XLSX export should succeed after review resolution.")
                assert_true(
                    ready_animal_xlsx.content[:4] == b"PK\x03\x04",
                    "Animal sheet XLSX export should be a ZIP-based workbook.",
                )
                animal_disposition = ready_animal_xlsx.headers.get("content-disposition", "")
                assert_true(
                    "filename*=UTF-8''" in animal_disposition and "animal sheet.xlsx" in animal_disposition,
                    "Animal sheet XLSX export should expose a safe fallback and UTF-8 filename.",
                )
                with zipfile.ZipFile(io.BytesIO(ready_animal_xlsx.content)) as workbook_zip:
                    assert_true(
                        "xl/workbook.xml" in workbook_zip.namelist()
                        and "xl/styles.xml" in workbook_zip.namelist()
                        and "xl/worksheets/sheet1.xml" in workbook_zip.namelist()
                        and "xl/worksheets/sheet2.xml" in workbook_zip.namelist(),
                        "Animal sheet XLSX should contain the required workbook parts.",
                    )
                    animal_workbook_xml = workbook_zip.read("xl/workbook.xml").decode("utf-8")
                    animal_sheet_xml = workbook_zip.read("xl/worksheets/sheet1.xml").decode("utf-8")
                    animal_trace_xml = workbook_zip.read("xl/worksheets/sheet2.xml").decode("utf-8")
                assert_true("Cage No." in animal_sheet_xml, "Animal sheet XLSX should include the template header.")
                assert_true("MT321" in animal_sheet_xml and "MT322" in animal_sheet_xml, "Animal sheet XLSX should include parent IDs.")
                assert_true("2026-05-02 10p" in animal_sheet_xml, "Animal sheet XLSX should include litter pup evidence.")
                assert_true("animal sheet" in animal_workbook_xml and "Export_Trace" in animal_workbook_xml, "Animal sheet XLSX should name workbook sheets clearly.")
                assert_true("<cols>" in animal_sheet_xml and 's="1"' in animal_sheet_xml, "Animal sheet XLSX should include column widths and styled headers.")
                assert_true("Source record" in animal_trace_xml, "Animal sheet XLSX should include traceability sheet.")
                animal_workbook = load_workbook(io.BytesIO(ready_animal_xlsx.content), data_only=True)
                assert_true("animal sheet" in animal_workbook.sheetnames, "openpyxl should load the animal sheet name.")
                assert_true("Export_Trace" in animal_workbook.sheetnames, "openpyxl should load the animal trace sheet.")
                animal_sheet = animal_workbook["animal sheet"]
                assert_true(
                    [animal_sheet.cell(1, column).value for column in range(1, 9)]
                    == ["Cage No.", "Strain", "Sex", "I.D", "genotype", "DOB", "Mating date", "Pubs"],
                    "openpyxl should read the animal sheet headers.",
                )
                animal_values = list(animal_sheet.iter_rows(min_row=2, values_only=True))
                assert_true(
                    any("MT321" in str(cell) for row in animal_values for cell in row if cell is not None)
                    and any("MT322" in str(cell) for row in animal_values for cell in row if cell is not None),
                    "openpyxl should read animal sheet parent IDs.",
                )
                assert_true(
                    any("2026-05-02 10p" in str(cell) for row in animal_values for cell in row if cell is not None),
                    "openpyxl should read animal sheet litter pup evidence.",
                )
                animal_trace_sheet = animal_workbook["Export_Trace"]
                assert_true(
                    animal_trace_sheet.cell(1, 3).value == "Source record"
                    and animal_trace_sheet.cell(1, 4).value == "Boundary",
                    "openpyxl should read animal trace headers.",
                )
                ready_logs = client.get("/api/export-log").json()
                assert_true(
                    any(item["export_type"] == "separation_xlsx" and item["status"] == "generated" for item in ready_logs),
                    "Export log should record generated separation XLSX exports.",
                )
                assert_true(
                    any(item["export_type"] == "animal_sheet_xlsx" and item["status"] == "generated" for item in ready_logs),
                    "Export log should record generated animal sheet XLSX exports.",
                )
                ready_export_log = next(item for item in ready_logs if item["export_type"] == "mouse_csv")
                assert_true(ready_export_log["status"] == "generated", "Ready export should create a generated export log entry.")
                assert_true(
                    ready_export_log["blocked_review_count"] == 0,
                    "Ready export log should record zero review blockers after resolution.",
                )

    print("Local app scaffold verification passed.")


if __name__ == "__main__":
    main()
