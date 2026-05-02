#!/usr/bin/env python3
"""Parse predecessor colony workbooks into reviewable JSON.

These Excel files are import/export views, not canonical state. The parser keeps
raw values and source-cell traceability so newer cage-card photos can override or
challenge workbook rows only through review.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as error:
    raise SystemExit(
        "openpyxl is required. Install it with `python -m pip install openpyxl` "
        "or run this script with the project virtual environment."
    ) from error


ANIMAL_HEADERS = {
    "cage no.": "cage_no",
    "cage no": "cage_no",
    "cage": "cage_no",
    "strain": "strain",
    "sex": "sex",
    "i.d": "display_id",
    "id": "display_id",
    "genotype": "genotype",
    "dob": "dob",
    "mating date": "mating_date",
    "pup": "pubs",
    "pubs": "pubs",
    "pups": "pubs",
}

SEPARATION_HEADERS = {
    "strain": "strain",
    "genotype": "genotype",
    "sex": "total",
    "total": "total",
    "dob": "dob",
    "wt": "wt",
    "tg": "tg",
    "sampling point": "sampling_point",
}


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return " ".join(str(value).split())


def normalize_header(value: Any) -> str:
    return normalize_cell(value).strip().lower().replace("i.d.", "i.d").replace("i.d", "i.d")


def detect_header(ws: Worksheet, expected: dict[str, str]) -> tuple[int, dict[str, int]]:
    best_row = 0
    best_map: dict[str, int] = {}
    for row_number in range(1, min(ws.max_row, 20) + 1):
        found: dict[str, int] = {}
        for col_number in range(1, ws.max_column + 1):
            key = normalize_header(ws.cell(row_number, col_number).value)
            if key in expected and expected[key] not in found:
                found[expected[key]] = col_number
        if len(found) > len(best_map):
            best_row = row_number
            best_map = found
    return best_row, best_map


def infer_sex_candidate(value: str) -> str:
    lowered = normalize_cell(value).lower()
    tokens = set(re.findall(r"[a-z]+", lowered))
    if lowered in {"m", "male"} or "m" in tokens or "male" in tokens or "\u2642" in value:
        return "male"
    if lowered in {"f", "female"} or "f" in tokens or "female" in tokens or "\u2640" in value:
        return "female"
    return ""


def compact_kind(sex: str, display_id: str = "", mating_date: str = "", pubs: str = "") -> str:
    lowered = normalize_cell(sex).lower()
    display = normalize_cell(display_id).lower()
    if (
        re.match(r"^f\d+$", lowered)
        or re.match(r"^\d+\s*p$", display)
        or pubs
        or "separated" in lowered
        or "dead" in lowered
    ):
        return "litter_or_offspring_snapshot"
    if display_id or infer_sex_candidate(sex):
        return "parent_or_mouse_snapshot"
    return "unclassified_workbook_row"


def parse_count(raw_total: str) -> int | None:
    match = re.search(r"(\d+)\s*p?", raw_total, re.IGNORECASE)
    return int(match.group(1)) if match else None


def parse_sex(raw_total: str) -> str:
    return infer_sex_candidate(raw_total)


def cell_value(ws: Worksheet, row: int, columns: dict[str, int], name: str) -> str:
    column = columns.get(name)
    return normalize_cell(ws.cell(row, column).value) if column else ""


def cell_ref(ws: Worksheet, row: int, columns: dict[str, int], name: str) -> str:
    column = columns.get(name)
    return ws.cell(row, column).coordinate if column else ""


def looks_like_strain_value(value: str) -> bool:
    text = normalize_cell(value)
    if not text:
        return False
    return not bool(re.match(r"^(c[- ]?)?\d+$", text, re.IGNORECASE))


def infer_animal_columns(ws: Worksheet, header_row: int, columns: dict[str, int]) -> dict[str, int]:
    inferred = dict(columns)
    display_column = inferred.get("display_id")
    dob_column = inferred.get("dob")
    if "genotype" not in inferred and display_column and dob_column and display_column + 1 < dob_column:
        inferred["genotype"] = display_column + 1
    return inferred


def parse_animal_sheet(ws: Worksheet, source_file: Path) -> dict[str, Any]:
    header_row, columns = detect_header(ws, ANIMAL_HEADERS)
    columns = infer_animal_columns(ws, header_row, columns)
    required = {"cage_no", "sex", "display_id", "dob"}
    if not required.issubset(columns):
        raise ValueError(f"Animal sheet headers not found in {ws.title!r}. Found: {sorted(columns)}")

    rows: list[dict[str, Any]] = []
    carried_cage = ""
    carried_strain = ""
    for row_number in range(header_row + 1, ws.max_row + 1):
        cage = cell_value(ws, row_number, columns, "cage_no")
        strain = cell_value(ws, row_number, columns, "strain")
        if "strain" not in columns and looks_like_strain_value(cage):
            strain = cage
            cage = ""
        if cage:
            carried_cage = cage
        if strain:
            carried_strain = strain
        elif not carried_strain:
            carried_strain = ws.title

        sex = cell_value(ws, row_number, columns, "sex")
        display_id = cell_value(ws, row_number, columns, "display_id")
        genotype = cell_value(ws, row_number, columns, "genotype")
        dob = cell_value(ws, row_number, columns, "dob")
        mating_date = cell_value(ws, row_number, columns, "mating_date")
        pubs = cell_value(ws, row_number, columns, "pubs")
        if not any([sex, display_id, genotype, dob, mating_date, pubs]):
            continue

        rows.append(
            {
                "row_type": compact_kind(sex, display_id=display_id, mating_date=mating_date, pubs=pubs),
                "source_layer": "export or view",
                "cage_no_raw": carried_cage,
                "strain_raw": carried_strain,
                "sex_raw": sex,
                "sex_candidate": infer_sex_candidate(sex),
                "display_id_raw": display_id,
                "genotype_raw": genotype,
                "dob_raw": dob,
                "mating_date_raw": mating_date,
                "pubs_raw": pubs,
                "review_status": "candidate",
                "source_sheet": ws.title,
                "source_row_number": row_number,
                "source_cells": {
                    "cage_no": cell_ref(ws, row_number, columns, "cage_no"),
                    "strain": cell_ref(ws, row_number, columns, "strain"),
                    "sex": cell_ref(ws, row_number, columns, "sex"),
                    "display_id": cell_ref(ws, row_number, columns, "display_id"),
                    "genotype": cell_ref(ws, row_number, columns, "genotype"),
                    "dob": cell_ref(ws, row_number, columns, "dob"),
                    "mating_date": cell_ref(ws, row_number, columns, "mating_date"),
                    "pubs": cell_ref(ws, row_number, columns, "pubs"),
                },
            }
        )

    return base_payload(source_file, ws, "legacy_animal_sheet", rows)


def parse_separation_sheet(ws: Worksheet, source_file: Path) -> dict[str, Any]:
    header_row, columns = detect_header(ws, SEPARATION_HEADERS)
    for col_number in range(1, ws.max_column + 1):
        sub_header = normalize_header(ws.cell(header_row + 1, col_number).value)
        if sub_header in {"wt", "tg"}:
            columns[sub_header] = col_number
    required = {"strain", "genotype", "total", "dob"}
    if not required.issubset(columns):
        raise ValueError(f"Separation sheet headers not found in {ws.title!r}. Found: {sorted(columns)}")

    rows: list[dict[str, Any]] = []
    carried_strain = ""
    carried_genotype = ""
    for row_number in range(header_row + 1, ws.max_row + 1):
        strain = cell_value(ws, row_number, columns, "strain")
        genotype = cell_value(ws, row_number, columns, "genotype")
        if strain:
            carried_strain = strain
        if genotype:
            carried_genotype = genotype

        total = cell_value(ws, row_number, columns, "total")
        dob = cell_value(ws, row_number, columns, "dob")
        wt = cell_value(ws, row_number, columns, "wt")
        tg = cell_value(ws, row_number, columns, "tg")
        sampling_point = cell_value(ws, row_number, columns, "sampling_point")
        if not any([total, dob]) and normalize_header(wt) == "wt" and normalize_header(tg) == "tg":
            continue
        if not any([total, dob, wt, tg]):
            continue

        rows.append(
            {
                "row_type": "separation_summary_row",
                "source_layer": "export or view",
                "strain_raw": carried_strain,
                "genotype_raw": carried_genotype,
                "total_raw": total,
                "sex_candidate": parse_sex(total),
                "count_candidate": parse_count(total),
                "dob_raw": dob,
                "wt_raw": wt,
                "tg_raw": tg,
                "sampling_point_raw": sampling_point,
                "review_status": "candidate",
                "source_sheet": ws.title,
                "source_row_number": row_number,
                "source_cells": {
                    "strain": cell_ref(ws, row_number, columns, "strain"),
                    "genotype": cell_ref(ws, row_number, columns, "genotype"),
                    "total": cell_ref(ws, row_number, columns, "total"),
                    "dob": cell_ref(ws, row_number, columns, "dob"),
                    "wt": cell_ref(ws, row_number, columns, "wt"),
                    "tg": cell_ref(ws, row_number, columns, "tg"),
                    "sampling_point": cell_ref(ws, row_number, columns, "sampling_point"),
                },
            }
        )

    return base_payload(source_file, ws, "legacy_separation_status", rows)


def base_payload(source_file: Path, ws: Worksheet, workbook_kind: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    sheet_names = sorted({str(row.get("source_sheet") or ws.title) for row in rows}) or [ws.title]
    return {
        "layer": "parsed or intermediate result",
        "source_layer": "export or view",
        "workbook_kind": workbook_kind,
        "source_file_name": source_file.name,
        "source_file_path": str(source_file),
        "sheet_name": sheet_names[0] if len(sheet_names) == 1 else ", ".join(sheet_names),
        "parsed_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "notes": [
            "Predecessor Excel rows are snapshots/views and must not overwrite newer cage-card photo evidence.",
            "Rows require review before becoming canonical cage, mouse, mating, litter, or genotype state.",
        ],
    }


def merge_payloads(source_file: Path, workbook_kind: str, payloads: list[dict[str, Any]]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for payload in payloads:
        rows.extend(payload.get("rows") or [])
    sheet_names = sorted({str(row.get("source_sheet") or "") for row in rows if row.get("source_sheet")})
    return {
        "layer": "parsed or intermediate result",
        "source_layer": "export or view",
        "workbook_kind": workbook_kind,
        "source_file_name": source_file.name,
        "source_file_path": str(source_file),
        "sheet_name": ", ".join(sheet_names),
        "parsed_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "notes": [
            "Predecessor Excel rows are snapshots/views and must not overwrite newer cage-card photo evidence.",
            "Rows require review before becoming canonical cage, mouse, mating, litter, or genotype state.",
            "Multiple workbook sheets were parsed into one reviewable import when their shapes matched the selected kind.",
        ],
    }


def parse_workbook(path: Path, kind: str = "auto", sheet_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
        if kind == "animal":
            return parse_animal_sheet(ws, path)
        if kind == "separation":
            return parse_separation_sheet(ws, path)
        animal_header, animal_columns = detect_header(ws, ANIMAL_HEADERS)
        separation_header, separation_columns = detect_header(ws, SEPARATION_HEADERS)
        if animal_header and len(animal_columns) >= len(separation_columns):
            return parse_animal_sheet(ws, path)
        if separation_header:
            return parse_separation_sheet(ws, path)
        raise ValueError(f"Could not detect supported legacy workbook shape for {path.name} sheet {sheet_name!r}.")

    if kind in {"animal", "separation"}:
        parser = parse_animal_sheet if kind == "animal" else parse_separation_sheet
        workbook_kind = "legacy_animal_sheet" if kind == "animal" else "legacy_separation_status"
        payloads: list[dict[str, Any]] = []
        errors: list[str] = []
        for ws in wb.worksheets:
            try:
                payload = parser(ws, path)
            except ValueError as error:
                errors.append(str(error))
                continue
            if payload.get("rows"):
                payloads.append(payload)
        if payloads:
            return merge_payloads(path, workbook_kind, payloads)
        raise ValueError("; ".join(errors) or f"No supported {kind} sheets found in {path.name}.")

    ws = wb[wb.sheetnames[0]]
    if kind == "animal":
        return parse_animal_sheet(ws, path)
    if kind == "separation":
        return parse_separation_sheet(ws, path)

    animal_header, animal_columns = detect_header(ws, ANIMAL_HEADERS)
    separation_header, separation_columns = detect_header(ws, SEPARATION_HEADERS)
    if animal_header and len(animal_columns) >= len(separation_columns):
        return parse_animal_sheet(ws, path)
    if separation_header:
        return parse_separation_sheet(ws, path)
    raise ValueError(f"Could not detect supported legacy workbook shape for {path.name}.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse predecessor colony workbook views into reviewable JSON.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--kind", choices=["auto", "animal", "separation"], default="auto")
    parser.add_argument("--sheet")
    parser.add_argument("--out", type=Path)
    args = parser.parse_args()

    payload = parse_workbook(args.workbook, kind=args.kind, sheet_name=args.sheet)
    output = json.dumps(payload, ensure_ascii=False, indent=2)
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(output + "\n", encoding="utf-8")
    else:
        print(output)


if __name__ == "__main__":
    main()
