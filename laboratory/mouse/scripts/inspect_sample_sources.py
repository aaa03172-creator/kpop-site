#!/usr/bin/env python3
"""Inspect local sample source files without changing them.

The output is intentionally compact so predecessor Excel files can be treated as
import/export views and cage-card photos can remain raw source evidence.
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from PIL import Image
except ModuleNotFoundError:
    Image = None


def normalize(value: Any, max_chars: int = 100) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())[:max_chars]


def inspect_workbook(path: Path, max_rows: int, max_cols: int) -> None:
    print(f"FILE\t{path.name}")
    wb = load_workbook(path, read_only=True, data_only=True)
    print("SHEETS\t" + ", ".join(wb.sheetnames))
    for ws in wb.worksheets:
        print(f"SHEET\t{ws.title}\trows={ws.max_row}\tcols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
            values = [normalize(value) for value in row[: min(ws.max_column, max_cols)]]
            print("ROW\t" + "\t".join(values))
        print("---")


def inspect_photo_folder(path: Path, pattern: str) -> None:
    print(f"PHOTO_FOLDER\t{path}")
    for photo in sorted(path.glob(pattern)):
        modified = datetime.fromtimestamp(photo.stat().st_mtime).isoformat(timespec="seconds")
        size = photo.stat().st_size
        if Image is None:
            print(f"PHOTO\t{photo.name}\tunknown\tmodified={modified}\tsize={size}")
        else:
            dimensions = "unreadable"
            try:
                with Image.open(photo) as image:
                    dimensions = f"{image.width}x{image.height}"
            except Exception:
                dimensions = "unreadable"
            print(
                f"PHOTO\t{photo.name}\t{dimensions}"
                f"\tmodified={modified}\tsize={size}"
            )


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect sample workbook/photo sources.")
    parser.add_argument("--workbook", action="append", type=Path, default=[])
    parser.add_argument("--photo-folder", type=Path)
    parser.add_argument("--photo-pattern", default="*.jpg")
    parser.add_argument("--max-rows", type=int, default=12)
    parser.add_argument("--max-cols", type=int, default=14)
    args = parser.parse_args()

    for workbook in args.workbook:
        inspect_workbook(workbook, args.max_rows, args.max_cols)
    if args.photo_folder:
        inspect_photo_folder(args.photo_folder, args.photo_pattern)


if __name__ == "__main__":
    main()
