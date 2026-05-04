#!/usr/bin/env python3
"""Parse periodic mouse mating distribution workbooks into reviewable JSON.

This script treats the workbook as raw source evidence. The JSON it produces is
parsed/intermediate data for assignment scope and strain-master suggestions; it
must not be treated as canonical cage/card state.
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as error:
    raise SystemExit(
        "openpyxl is required. Install it with `python -m pip install openpyxl` "
        "or run this script with the bundled workspace Python."
    ) from error


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return " ".join(str(value).split())


def parse_int_like(value: Any) -> int | None:
    text = normalize_cell(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def detect_blocks(ws: Worksheet) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for col in range(1, ws.max_column - 2):
        group = normalize_cell(ws.cell(1, col).value)
        mating_header = normalize_cell(ws.cell(1, col + 1).value)
        cage_header = normalize_cell(ws.cell(1, col + 2).value)
        mating_cage_header = normalize_cell(ws.cell(1, col + 3).value)
        if (
            group
            and "mating" in mating_header.lower()
            and "cage" in cage_header.lower()
            and "mating" in mating_cage_header.lower()
        ):
            blocks.append(
                {
                    "institution_or_group": group,
                    "person_col": col,
                    "mating_col": col + 1,
                    "cage_count_col": col + 2,
                    "mating_cage_count_col": col + 3,
                }
            )
    return blocks


def parse_distribution_workbook(workbook_path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        blocks = detect_blocks(ws)
        if not blocks:
            raise ValueError("No distribution column blocks found in header row.")

        rows: list[dict[str, Any]] = []
        carried_people: dict[int, str] = {}
        for row_index in range(2, ws.max_row + 1):
            for block in blocks:
                person_col = block["person_col"]
                raw_person = normalize_cell(ws.cell(row_index, person_col).value)
                if raw_person:
                    carried_people[person_col] = raw_person
                responsible = carried_people.get(person_col, "")
                mating_type = normalize_cell(ws.cell(row_index, block["mating_col"]).value)
                cage_count = normalize_cell(ws.cell(row_index, block["cage_count_col"]).value)
                mating_cage_count = normalize_cell(ws.cell(row_index, block["mating_cage_count_col"]).value)
                if not any([responsible, mating_type, cage_count, mating_cage_count]):
                    continue
                if not mating_type and not cage_count and not mating_cage_count:
                    continue
                rows.append(
                    {
                        "institution_or_group": block["institution_or_group"],
                        "responsible_person_raw": responsible,
                        "mating_type_raw": mating_type,
                        "cage_count_raw": cage_count,
                        "cage_count_value": parse_int_like(cage_count),
                        "mating_cage_count_raw": mating_cage_count,
                        "mating_cage_count_value": parse_int_like(mating_cage_count),
                        "source_sheet": ws.title,
                        "source_row_number": row_index,
                        "source_cells": {
                            "responsible_person": ws.cell(row_index, person_col).coordinate,
                            "mating_type": ws.cell(row_index, block["mating_col"]).coordinate,
                            "cage_count": ws.cell(row_index, block["cage_count_col"]).coordinate,
                            "mating_cage_count": ws.cell(row_index, block["mating_cage_count_col"]).coordinate,
                        },
                        "review_status": "candidate",
                    }
                )

        return {
            "layer": "parsed or intermediate result",
            "source_file_name": workbook_path.name,
            "source_file_path": str(workbook_path),
            "sheet_name": ws.title,
            "received_date": None,
            "parsed_at": datetime.now().isoformat(timespec="seconds"),
            "detected_blocks": [
                {
                    "institution_or_group": block["institution_or_group"],
                    "columns": {
                        "responsible_person": block["person_col"],
                        "mating_type": block["mating_col"],
                        "cage_count": block["cage_count_col"],
                        "mating_cage_count": block["mating_cage_count_col"],
                    },
                }
                for block in blocks
            ],
            "rows": rows,
        }
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse a mouse mating distribution workbook into JSON.")
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx distribution workbook.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the first sheet.")
    parser.add_argument("--out", type=Path, help="Output JSON path. Defaults to stdout.")
    args = parser.parse_args()

    parsed = parse_distribution_workbook(args.workbook, args.sheet)
    output = json.dumps(parsed, ensure_ascii=False, indent=2)
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(output + "\n", encoding="utf-8")
    else:
        print(output)


if __name__ == "__main__":
    main()
